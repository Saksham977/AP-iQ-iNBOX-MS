"""
Payment Category: match invoice/payment details from a screenshot, pasted
email/text, PDF, or Excel file against the Vendor Ledger (the same BC export
used by "Run Bulk Status"), and produce a bulk-status-style results workbook.

Matching key: invoice number, exact match (case-insensitive, trimmed) against
the ledger. Confidence is derived from how well the OTHER extracted fields
(vendor / amount / date) agree with the matched ledger row -- not from asking
Claude to self-report confidence -- so it also works for the two input types
(PDF, Excel) that never touch Claude at all.

Payment status (the "Status" column) is derived purely from two ledger
figures -- Remaining Amt. (LCY) and Remaining Amount Sales Invoice -- per the
house payment-eligibility rules (see _derive_payment_status below). It is
NOT the same status computed by the Bulk Status skill (match_status.py),
which only ever sees Open/Due Date, not a customer-side sales figure.
"""
from collections import namedtuple
from datetime import datetime, date, timedelta

import openpyxl
import pdfplumber

# Returned by load_ledger() — rows list plus O(1) dict indexes.
LedgerIndex = namedtuple("LedgerIndex", ["rows", "by_invoice", "by_rfp"])

# ---------------------------------------------------------------------------
# Flexible column detection -- ledger/PDF/Excel headers vary, so we search
# for likely keywords rather than hardcoding exact header text. Each spec is
# {"hints": [...]} checked in order across ALL headers (so a more specific
# earlier hint always wins over a later, more generic one regardless of
# column position), with an optional "exclude" list of substrings that
# disqualify an otherwise-matching header (used to keep look-alike columns,
# e.g. "Remaining Amount Sales Invoice", from being grabbed by a more
# generic hint meant for a different column, e.g. "Remaining Amount").
# ---------------------------------------------------------------------------

COLUMN_SPECS = {
    "vendor": {"hints": ["vendor name", "supplier", "payee", "vendor"]},
    # "external document no" is Business Central's actual header for the
    # vendor's invoice number on a raw Vendor Ledger Entries export (see
    # skills_mirror/bulk-status/scripts/match_status.py) -- without this,
    # matching against a real BC export silently found zero invoice numbers.
    "invoice_no": {"hints": ["external document no", "invoice", "inv no", "inv. no"]},
    "rfp_no": {"hints": ["rfp"]},
    "amount": {"hints": ["remaining amount", "amount", "total", "value"],
               "exclude": ["sales invoice", "(lcy)"]},
    # "date" now means the invoice's posting/document date, not the due
    # date -- "Due Date" is its own column below. Excluding "due date" stops
    # the generic "date" hint from grabbing the Due Date column when no
    # dedicated posting/document date header exists.
    "date": {"hints": ["posting date", "document date", "invoice date", "date"],
              "exclude": ["due date"]},
    "due_date": {"hints": ["due date"]},
    "status": {"hints": ["status update", "status"]},
    "open": {"hints": ["open"]},
    "reference": {"hints": ["entry no", "entry number", "ledger ref", "document no", "doc no"]},
    "remaining_amt_lcy": {"hints": ["remaining amt. (lcy)", "remaining amt (lcy)"]},
    "remaining_sales_amt": {"hints": ["remaining amount sales invoice"]},
}


def _find_column(headers, spec):
    """See COLUMN_SPECS docstring above for the hints/exclude semantics."""
    hints = spec.get("hints", [])
    exclude = spec.get("exclude", [])
    lower_headers = [str(h).strip().lower() if h else "" for h in headers]
    for kw in hints:
        for idx, h_low in enumerate(lower_headers):
            if kw in h_low and not any(ex in h_low for ex in exclude):
                return idx
    return None


def _normalize(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _to_number(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


# ---------------------------------------------------------------------------
# Payment status -- house rules, computed ONLY from Remaining Amt. (LCY) and
# Remaining Amount Sales Invoice (+ Due Date for timing). See the "Business
# Logic" spec: rules 1-5.
# ---------------------------------------------------------------------------

STATUS_PAID = "Paid"
STATUS_CUSTOMER_NOT_PAID = "Customer Not Paid"
STATUS_PAY_NEXT_CYCLE = "Ready for Payment – Pay in Next Payment Cycle"
STATUS_NOT_DUE_YET = "Ready for Payment – Not Due Yet"
STATUS_DUE_DATE_UNKNOWN = "Ready for Payment – Due Date Unknown"
# Rare overlap case (Remaining Amt. (LCY) = 0 AND Remaining Amount Sales
# Invoice != 0 at the same time) -- flagged distinctly rather than silently
# resolved either way, per explicit instruction to treat it separately.
STATUS_PAID_BUT_CUSTOMER_NOT_PAID = "Paid, but Customer Not Paid (rare – please verify)"
STATUS_UNKNOWN = "Unknown – insufficient ledger data"

PAYMENT_TIMING_WINDOW_DAYS = 2


def _derive_payment_status(remaining_lcy, remaining_sales, due_date, today=None):
    if today is None:
        today = date.today()

    lcy = _to_number(remaining_lcy)
    sales = _to_number(remaining_sales)

    if lcy is None and sales is None:
        return STATUS_UNKNOWN

    lcy_zero = lcy is not None and lcy == 0
    lcy_nonzero = lcy is not None and lcy != 0
    sales_zero = sales is not None and sales == 0
    sales_nonzero = sales is not None and sales != 0

    # Rule: rare overlap -- LCY says paid, but the customer-side figure says
    # otherwise. Called out separately rather than picking one silently.
    if lcy_zero and sales_nonzero:
        return STATUS_PAID_BUT_CUSTOMER_NOT_PAID

    # Rule 1: Invoice already paid.
    if lcy_zero:
        return STATUS_PAID

    # Rule 2: Customer has not paid yet.
    if sales_nonzero:
        return STATUS_CUSTOMER_NOT_PAID

    # Rule 3: Ready for payment.
    if sales_zero and lcy_nonzero:
        due = _as_date(due_date)
        if due is None:
            return STATUS_DUE_DATE_UNKNOWN
        # Rule 4: due date passed, or within the payment-timing window.
        if due <= today + timedelta(days=PAYMENT_TIMING_WINDOW_DAYS):
            return STATUS_PAY_NEXT_CYCLE
        return STATUS_NOT_DUE_YET

    return STATUS_UNKNOWN


def is_ready_for_payment(remaining_lcy, remaining_sales):
    """Rule 5: Due/Overdue determinations must be based ONLY on this
    condition -- Remaining Amount Sales Invoice = 0 AND
    Remaining Amt. (LCY) != 0 -- used by Payment Insights, not just Status."""
    lcy = _to_number(remaining_lcy)
    sales = _to_number(remaining_sales)
    return sales is not None and sales == 0 and lcy is not None and lcy != 0


def _load_rows_from_sheet(headers, rows):
    """Map a header row + data rows into normalized dicts using COLUMN_SPECS."""
    col_idx = {key: _find_column(headers, spec) for key, spec in COLUMN_SPECS.items()}
    # A header can only mean one thing -- if "reference" landed on the same
    # column as "invoice_no" (e.g. both matched on "document no" inside
    # "External Document No."), drop the reference match rather than
    # silently duplicating the invoice number as the ledger reference too.
    if col_idx.get("reference") is not None and col_idx.get("reference") == col_idx.get("invoice_no"):
        col_idx["reference"] = None

    records = []
    for row_num, row in enumerate(rows, start=2):  # header is row 1
        raw = {}
        for h_idx, h in enumerate(headers):
            if h is None:
                continue
            raw[str(h).strip()] = row[h_idx] if h_idx < len(row) else None

        def get(key):
            idx = col_idx.get(key)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        invoice_no = _normalize(get("invoice_no"))
        if not invoice_no:
            continue

        raw_due_date = get("due_date")
        raw_lcy = get("remaining_amt_lcy")
        raw_sales = get("remaining_sales_amt")
        status = _derive_payment_status(raw_lcy, raw_sales, raw_due_date)

        records.append({
            "vendor": _normalize(get("vendor")),
            "invoice_no": invoice_no,
            "rfp_no": _normalize(get("rfp_no")),
            "amount": get("amount"),
            "date": _normalize(get("date")),
            "due_date": _normalize(raw_due_date),
            "due_date_raw": raw_due_date,
            "status": status,
            "reference": _normalize(get("reference")) or f"Row {row_num}",
            "remaining_amt_lcy": raw_lcy,
            "remaining_sales_amt": raw_sales,
            "raw": raw,
            "row_num": row_num,
        })
    return records


# ---------------------------------------------------------------------------
# Ledger loading
# ---------------------------------------------------------------------------

def load_ledger_from_rows(headers, data_rows):
    """Build a LedgerIndex from already-read raw rows — no file I/O.
    Called by the unified ledger loader in server.py so the file is opened once."""
    rows = _load_rows_from_sheet(headers, data_rows)
    by_invoice = {r["invoice_no"].upper(): r for r in rows if r["invoice_no"]}
    by_rfp = {r["rfp_no"].upper(): r for r in rows if r["rfp_no"]}
    return LedgerIndex(rows=rows, by_invoice=by_invoice, by_rfp=by_rfp)


def load_ledger(ledger_path):
    """Load the BC Vendor Ledger and return a LedgerIndex with O(1) lookup dicts."""
    wb = openpyxl.load_workbook(ledger_path, data_only=True, read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        return LedgerIndex(rows=[], by_invoice={}, by_rfp={})
    headers, data_rows = all_rows[0], all_rows[1:]
    return load_ledger_from_rows(headers, data_rows)


def match_by_invoice_no(ledger_index, invoice_no):
    if not invoice_no:
        return None
    return ledger_index.by_invoice.get(_normalize(invoice_no).upper())


def match_by_rfp_no(ledger_index, rfp_no):
    if not rfp_no:
        return None
    return ledger_index.by_rfp.get(_normalize(rfp_no).upper())


# ---------------------------------------------------------------------------
# Direct extraction: PDF / Excel (no Claude)
# ---------------------------------------------------------------------------

def extract_from_excel(file_path):
    """Read a row-per-invoice Excel file directly -- no AI involved."""
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []
    headers, data_rows = all_rows[0], all_rows[1:]
    return _load_rows_from_sheet(headers, data_rows)


def extract_from_pdf(file_path):
    """
    Read a structured/tabular PDF (an exported invoice list or ledger
    extract) directly -- no AI involved. Extracts every table found across
    all pages and maps rows the same way as the Excel path.
    """
    records = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                if not table or len(table) < 2:
                    continue
                headers, data_rows = table[0], table[1:]
                records.extend(_load_rows_from_sheet(headers, data_rows))
    return records


# ---------------------------------------------------------------------------
# Confidence scoring: derived from match quality, not self-reported by Claude
# ---------------------------------------------------------------------------

def _amounts_close(a, b, tolerance=0.01):
    try:
        return abs(float(a) - float(b)) <= tolerance
    except (TypeError, ValueError):
        return False


def score_confidence(extracted, ledger_row):
    """
    extracted: dict with optional vendor/amount/date (as parsed/extracted
    from the source, before ledger lookup).
    ledger_row: the matched ledger row dict, or None if no match was found.
    Returns (score 0-100, label).
    """
    if ledger_row is None:
        return 0, "No match"

    score = 60  # baseline: invoice number matched exactly
    checks = 0

    vendor_ex = (extracted.get("vendor") or "").strip().lower()
    if vendor_ex:
        checks += 1
        if vendor_ex in (ledger_row["vendor"] or "").lower() or (ledger_row["vendor"] or "").lower() in vendor_ex:
            score += 15

    amount_ex = extracted.get("amount")
    if amount_ex is not None and amount_ex != "":
        checks += 1
        if _amounts_close(amount_ex, ledger_row["amount"]):
            score += 15

    date_ex = (extracted.get("date") or "").strip()
    if date_ex:
        checks += 1
        if date_ex[:10] == (ledger_row["date"] or "")[:10]:
            score += 10

    if checks == 0:
        # Only the invoice number was available to check -- exact key match
        # on its own is a reasonably strong signal, but nothing to corroborate.
        score = 75

    score = min(score, 100)
    label = "High" if score >= 85 else "Medium" if score >= 60 else "Low"
    return score, label


# ---------------------------------------------------------------------------
# Assembling and exporting results
# ---------------------------------------------------------------------------

CORE_COLUMNS = [
    "Vendor", "Invoice No.", "RFP No.", "Amount", "Date", "Due Date", "Status",
    "Ledger Ref.", "Match", "Confidence", "Source",
    "Remaining Amt. (LCY)", "Remaining Amount Sales Invoice",
]
EXTRA_META_COLUMNS = ["Confidence Score"]


def _is_covered_header(header):
    """True if this raw ledger header is already represented by one of the
    core/meta output columns (via the same keyword matching used to find it),
    so it isn't duplicated as an "extra" column."""
    h_low = str(header or "").strip().lower()
    for spec in COLUMN_SPECS.values():
        if any(kw in h_low for kw in spec.get("hints", [])):
            return True
    return False


def build_record(source_type, extracted, ledger_row):
    """
    Combine one extracted/parsed record with its ledger match (if any) into
    the unified output shape. Ledger values win where available (they're the
    authoritative source); extracted values are the fallback for unmatched
    invoices so the row isn't blank.
    """
    score, label = score_confidence(extracted, ledger_row)
    matched = ledger_row is not None

    extra_fields = {}
    if matched:
        for key, value in (ledger_row.get("raw") or {}).items():
            if not _is_covered_header(key):
                extra_fields[key] = value

    return {
        "Vendor": (ledger_row["vendor"] if matched else extracted.get("vendor")) or "",
        "Invoice No.": (ledger_row["invoice_no"] if matched else extracted.get("invoice_no")) or "",
        "RFP No.": (ledger_row["rfp_no"] if matched else extracted.get("rfp_no")) or "",
        "Amount": ledger_row["amount"] if matched else extracted.get("amount"),
        "Date": (ledger_row["date"] if matched else extracted.get("date")) or "",
        "Due Date": ledger_row["due_date"] if matched else "",
        "Status": (ledger_row["status"] if matched else "") or ("Unmatched" if not matched else ""),
        "Ledger Ref.": ledger_row["reference"] if matched else "",
        "Match": "Yes" if matched else "No",
        "Confidence": label,
        "Source": source_type,
        "Remaining Amt. (LCY)": ledger_row["remaining_amt_lcy"] if matched else None,
        "Remaining Amount Sales Invoice": ledger_row["remaining_sales_amt"] if matched else None,
        "Confidence Score": score,
        "_extra": extra_fields,
    }


def write_results_excel(records, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Category Results"

    extra_keys = []
    for r in records:
        for k in r.get("_extra", {}):
            if k not in extra_keys:
                extra_keys.append(k)

    headers = CORE_COLUMNS + EXTRA_META_COLUMNS + extra_keys
    ws.append(headers)

    for r in records:
        row = [r.get(col, "") for col in CORE_COLUMNS + EXTRA_META_COLUMNS]
        row += [r.get("_extra", {}).get(k, "") for k in extra_keys]
        ws.append(row)

    wb.save(out_path)
    return out_path
