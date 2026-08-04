#!/usr/bin/env python3
"""
Bulk Status matcher — compares a vendor/venue invoice-and-RFP overview against
a Business Central "Vendor Ledger Entries" export, and produces a status-update
workbook per the Bulk Status SOP.

Matching logic (from SOP, generalized):
  1. Direct match:  vendor's invoice number  ==  BC "External Document No."
  2. RFP match:     RFP number (own column, or last digit-run of a
                     description/reference column)  ==  BC "RFP No."
  3. Cross-check both directions; if they disagree, flag as a Discrepancy
     instead of silently trusting one side.
  4. If matched, pull BC's Remaining Amount / Due Date / Open flag and derive
     a plain-language payment status.
  5. If neither lookup finds anything in BC, flag as Not Found.

Usage:
    python match_status.py <bc_export.xlsx> <vendor_file.xlsx> <output_dir> [--today YYYY-MM-DD]

Output filename is always "Status Overview - <Vendor Name> - <Vendor No>.xlsx"
(house naming convention) — derived automatically from the Business Central
Vendor No./Vendor Name of the matched rows, not something you pass in.

Column detection is header-name based (case-insensitive substring match), so
this works on any BC export / vendor file pair that uses similar headers —
it is not hardcoded to a specific venue.
"""
import argparse
import os
import re
import sys
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Column detection helpers
# ---------------------------------------------------------------------------

BC_COLUMN_HINTS = {
    "ext_doc_no": {"hints": ["external document no"]},
    "doc_no": {"hints": ["document no"]},
    "rfp_no": {"hints": ["rfp no", "rfp number", "rfp id"]},
    "vendor_no": {"hints": ["vendor no"]},
    "vendor_name": {"hints": ["vendor name"]},
    "description": {"hints": ["description"]},
    # exclude "sales invoice" so this never accidentally grabs the sales-invoice
    # remaining-amount column when both are present in the same export.
    "remaining_amount": {"hints": ["remaining amount"], "exclude": ["sales invoice"]},
    # Not required — only present when the BC export includes the sales column.
    "remaining_sales_amt": {"hints": ["remaining amount sales invoice"], "required": False},
    "due_date": {"hints": ["due date"]},
    "open": {"hints": ["open"]},
    "currency": {"hints": ["currency code"]},
}

VENDOR_COLUMN_HINTS = {
    # Ordered most-specific → least-specific so a broader hint never steals a match
    # from a more precise one at the same position.
    "invoice_no": [
        # Dutch
        "factuurnummer", "factuur nr", "factuur no", "factuur id",
        # English multi-word
        "invoice number", "invoice num", "invoice no", "invoice #", "invoice id",
        "inv number", "inv num", "inv no", "inv #",
        "bill number", "bill num", "bill no", "bill #",
        "document number", "document no", "doc number", "doc no", "doc id",
        "reference number", "reference no", "ref number", "ref no", "ref #",
        # German
        "rechnungsnummer", "rechnung nr", "rechnung no",
        # French
        "numero de facture", "numéro de facture", "num facture",
        # Spanish / Portuguese
        "numero de factura", "núm. factura", "num factura",
        # Italian
        "numero fattura",
        # Single-word / short (placed last — broader match, higher false-positive risk)
        "factuur",    # Dutch: invoice
        "factura",    # Spanish / Portuguese
        "facture",    # French
        "fattura",    # Italian
        "rechnung",   # German
        "kwitantie",  # Dutch: receipt
        "bill",
    ],
    "description": [
        "factuuromschrijving", "invoice description", "description",
        "omschrijving",               # Dutch: description
        "opmerking",                  # Dutch: remark / note
        "memo", "notes", "note", "remark", "remarks", "details",
        "document", "text", "narrative",
        "bemerkung", "beschreibung",  # German
        "commentaire",                # French
    ],
    "rfp_no": [
        "rfp id", "rfp number", "rfp num", "rfp no", "rfp #",
        "po number", "po num", "po no", "po #",
        "event id", "event number", "event no",
        "booking id", "booking ref", "booking number",
        "meeting id", "project id", "project no", "project number",
        "order id", "order number", "order no",
        "rfp",  # single-word last — also matches "RFP ID", "RFP Number", etc.
    ],
}


def find_column(headers, hints, required=True, label="", exclude=None):
    """Return the 0-based index of the first header matching any hint substring.
    exclude: optional list of substrings that disqualify an otherwise-matching header."""
    lower_headers = [str(h).strip().lower() if h else "" for h in headers]
    for hint in hints:
        for i, h in enumerate(lower_headers):
            if hint in h:
                if exclude and any(ex in h for ex in exclude):
                    continue
                return i
    if required:
        raise ValueError(
            f"Could not find a column for '{label}' (looked for any of {hints}) "
            f"among headers: {headers}"
        )
    return None


def detect_columns(headers, hint_map, labels_required):
    cols = {}
    for key, spec in hint_map.items():
        if isinstance(spec, dict):
            hints = spec["hints"]
            exclude = spec.get("exclude")
            req = spec.get("required", key in labels_required)
        else:
            hints = spec
            exclude = None
            req = key in labels_required
        cols[key] = find_column(headers, hints, required=req, label=key, exclude=exclude)
    return cols


def _guess_invoice_col_by_values(sample_rows, n_cols):
    """Scan sample data rows and return the 0-based column index most likely to
    contain invoice numbers — typically short mixed-alphanumeric strings.
    Returns None when no strong candidate is found."""
    if not sample_rows or n_cols == 0:
        return None

    scores = [0] * n_cols
    for row in sample_rows:
        for col_idx in range(min(n_cols, len(row))):
            val = row[col_idx]
            if val is None:
                continue
            if isinstance(val, (datetime, date)):
                continue  # date column — not an invoice number
            if isinstance(val, (int, float)):
                s = str(int(val)) if isinstance(val, float) and val == int(val) else str(val)
                # Short pure-numeric strings are weak invoice signals
                if s.isdigit() and 3 <= len(s) <= 10:
                    scores[col_idx] += 1
                continue
            s = str(val).strip()
            if not s:
                continue
            has_alpha = any(c.isalpha() for c in s)
            has_digit = any(c.isdigit() for c in s)
            # Mixed alphanumeric (e.g. "INV-001", "F2024001") → strongest signal
            if has_alpha and has_digit and 3 <= len(s) <= 30:
                scores[col_idx] += 3
            elif has_alpha and 2 <= len(s) <= 25 and " " not in s:
                # Pure alpha without spaces (e.g. short codes) — weak signal
                scores[col_idx] += 1

    if not any(scores):
        return None
    best = max(range(n_cols), key=lambda i: scores[i])
    return best if scores[best] > 0 else None


def read_vendor_file_info(vendor_path, n_sample=15):
    """Open a vendor Excel file and return (headers, sample_rows, header_row_1based).
    Used by the server to get column headers for AI-based identification without
    re-parsing the full file inside process()."""
    vwb = openpyxl.load_workbook(vendor_path, data_only=True, read_only=True)
    vws = vwb.worksheets[0]
    v_header_row = find_header_row(vws, VENDOR_COLUMN_HINTS)
    headers = [c.value for c in next(vws.iter_rows(min_row=v_header_row, max_row=v_header_row))]
    sample_rows = list(vws.iter_rows(
        min_row=v_header_row + 1, max_row=v_header_row + n_sample, values_only=True
    ))
    vwb.close()
    return headers, sample_rows, v_header_row


def find_header_row(ws, hints_map, max_scan=30):
    """Scan the first max_scan rows and return the 1-based row index of the first
    row that contains at least one of the column hint substrings.  Falls back to
    row 1 when nothing matches (keeps original behaviour for well-formed exports)."""
    # hints_map values may be either a list of strings (VENDOR_COLUMN_HINTS) or a
    # dict with a "hints" key (BC_COLUMN_HINTS) — handle both.
    all_hints = []
    for spec in hints_map.values():
        if isinstance(spec, dict):
            all_hints.extend(spec.get("hints", []))
        else:
            all_hints.extend(spec)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        lower_vals = [str(v).strip().lower() if v is not None else "" for v in row]
        for hint in all_hints:
            if any(hint in v for v in lower_vals):
                return row_idx
    return 1


# ---------------------------------------------------------------------------
# RFP extraction
# ---------------------------------------------------------------------------

TRAILING_DIGITS = re.compile(r"(\d{4,})\s*$")  # trailing run of 4+ digits


def extract_rfp(text):
    """Pull the trailing digit run out of a free-text description, e.g.
    'LLY-RS-047712 MEETINGSELECT B.V. RFP ID: 759257' -> '759257'."""
    if not text:
        return None
    m = TRAILING_DIGITS.search(str(text).strip())
    return m.group(1) if m else None


def norm(v):
    """Normalize a lookup key: strip, uppercase not applied (case-sensitive doc numbers),
    but tolerate ints/floats coming back from Excel as e.g. 759257.0."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


# ---------------------------------------------------------------------------
# Core
# ---------------------------------------------------------------------------

STATUS_NOT_FOUND = "Not found in our records"
STATUS_PAID = "Paid"
STATUS_OVERDUE = "Overdue \u2013 payment pending"
STATUS_SCHEDULED = "Open \u2013 not yet due"
STATUS_DISCREPANCY_SUFFIX = " (discrepancy \u2013 please verify)"

INVALID_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]')


def sanitize_filename_part(text):
    text = "" if text is None else str(text).strip()
    text = INVALID_FILENAME_CHARS.sub("-", text)
    return text or "Unknown"


def build_output_filename(vendor_name, vendor_no):
    """House naming convention (always applies): 'Status Overview - <Vendor Name> - <Vendor No>.xlsx'"""
    name_part = sanitize_filename_part(vendor_name) if vendor_name else "Unknown Vendor"
    no_part = sanitize_filename_part(vendor_no) if vendor_no else "N-A"
    return f"Status Overview - {name_part} - {no_part}.xlsx"


def build_bc_dicts_from_rows(headers, data_rows):
    """Build (by_ext, by_rfp) dicts from already-read raw rows — no file I/O.
    Called by the unified ledger loader in server.py so the file is opened once."""
    cols = detect_columns(
        headers,
        BC_COLUMN_HINTS,
        labels_required={"ext_doc_no", "rfp_no", "remaining_amount", "due_date", "open"},
    )
    by_ext = {}
    by_rfp = {}
    for row in data_rows:
        if row is None:
            continue
        ext = norm(row[cols["ext_doc_no"]]) if cols["ext_doc_no"] is not None else None
        rfp = norm(row[cols["rfp_no"]]) if cols["rfp_no"] is not None else None
        rec = {
            "ext_doc_no": ext,
            "rfp_no": rfp,
            "doc_no": row[cols["doc_no"]] if cols.get("doc_no") is not None else None,
            "vendor_no": row[cols["vendor_no"]] if cols.get("vendor_no") is not None else None,
            "vendor_name": row[cols["vendor_name"]] if cols.get("vendor_name") is not None else None,
            "remaining_amount": row[cols["remaining_amount"]],
            "remaining_sales_amt": row[cols["remaining_sales_amt"]] if cols.get("remaining_sales_amt") is not None else None,
            "due_date": row[cols["due_date"]],
            "open": row[cols["open"]],
        }
        if ext:
            by_ext[ext] = rec
        if rfp:
            by_rfp.setdefault(rfp, rec)
    return by_ext, by_rfp


def load_bc_ledger(bc_path):
    wb = openpyxl.load_workbook(bc_path, data_only=True, read_only=True)
    # BC Vendor Ledger Entries exports always have headers in row 1.
    ws = wb.worksheets[0]
    for sheet in wb.worksheets:
        if "vendor ledger" in sheet.title.lower():
            ws = sheet
            break
    header_row = find_header_row(ws, BC_COLUMN_HINTS, max_scan=1)
    all_rows = list(ws.iter_rows(min_row=header_row, values_only=True))
    wb.close()
    if not all_rows:
        return {}, {}
    headers, data_rows = all_rows[0], all_rows[1:]
    return build_bc_dicts_from_rows(headers, data_rows)


def derive_payment_status(rec, today):
    if rec is None:
        return STATUS_NOT_FOUND
    open_flag = rec.get("open")
    is_open = bool(open_flag) and str(open_flag).strip() not in ("0", "False", "", "No")
    if not is_open:
        return STATUS_PAID
    due = rec.get("due_date")
    due_date = None
    if isinstance(due, datetime):
        due_date = due.date()
    elif isinstance(due, date):
        due_date = due
    if due_date and due_date < today:
        return STATUS_OVERDUE
    return STATUS_SCHEDULED


def process(bc_path, vendor_path, output_dir, today=None,
            _by_ext=None, _by_rfp=None, _vcols_override=None):
    """Process a vendor Excel file against the BC ledger.

    _vcols_override: optional dict mapping key → 0-based column index, produced
    by an AI column-identification step in the caller. Values here take
    precedence over hint-based detection and the value-pattern heuristic.
    """
    if today is None:
        today = date.today()

    if _by_ext is not None and _by_rfp is not None:
        by_ext, by_rfp = _by_ext, _by_rfp
    else:
        by_ext, by_rfp = load_bc_ledger(bc_path)

    vwb = openpyxl.load_workbook(vendor_path, data_only=True)
    vws = vwb.worksheets[0]
    v_header_row = find_header_row(vws, VENDOR_COLUMN_HINTS)
    headers = [c.value for c in next(vws.iter_rows(min_row=v_header_row, max_row=v_header_row))]

    # Layer 1: hint-based detection (no column is strictly required here so we
    # can fall through to the next layers instead of raising immediately).
    vcols = detect_columns(headers, VENDOR_COLUMN_HINTS, labels_required=set())

    # Layer 2: caller-supplied AI override (applied on top of hint results)
    if _vcols_override:
        for k, v in _vcols_override.items():
            if v is not None:
                vcols[k] = v

    # Layer 3: value-pattern heuristic — scan actual cell values when invoice_no
    # column still can't be identified from headers.
    if vcols.get("invoice_no") is None:
        sample_rows = list(vws.iter_rows(
            min_row=v_header_row + 1, max_row=v_header_row + 20, values_only=True
        ))
        vcols["invoice_no"] = _guess_invoice_col_by_values(sample_rows, len(headers))

    if vcols.get("invoice_no") is None:
        raise ValueError(
            f"Cannot identify the invoice-number column in '{os.path.basename(vendor_path)}'. "
            f"Headers found: {[h for h in headers if h is not None]}. "
            f"Add a column named 'Invoice No.', 'Factuur', 'Factuurnummer', or similar, "
            f"or rename an existing column to match."
        )

    has_rfp_col = vcols.get("rfp_no") is not None
    has_desc_col = vcols.get("description") is not None
    if not has_rfp_col and not has_desc_col:
        raise ValueError(
            "Vendor file has neither a dedicated RFP column nor a description column "
            "to extract an RFP number from — cannot cross-check."
        )

    out_headers = list(headers) + [
        "RFP No. (extracted)",
        "BC: External Doc. No. (via RFP)",
        "Match check 1 (invoice \u2194 RFP)",
        "BC: RFP No. (via invoice no.)",
        "Match check 2 (RFP \u2194 invoice)",
        "Match status",
        "BC: Remaining Amount",
        "BC: Remaining Amt. Sales Invoice",
        "BC: Due Date",
        "Status update",
    ]

    rows_out = []
    stats = {"matched": 0, "discrepancy": 0, "not_found": 0}
    vendor_counter = {}

    for row in vws.iter_rows(min_row=v_header_row + 1, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        invoice_no = norm(row[vcols["invoice_no"]])
        if invoice_no is None:
            # No invoice/document key on this row (e.g. a totals row) — skip it,
            # it isn't a real invoice line to report status on.
            continue
        rfp_extracted = None
        if has_rfp_col:
            rfp_extracted = norm(row[vcols["rfp_no"]])
        if not rfp_extracted and has_desc_col:
            rfp_extracted = extract_rfp(row[vcols["description"]])

        rec_by_invoice = by_ext.get(invoice_no) if invoice_no else None
        rec_by_rfp = by_rfp.get(rfp_extracted) if rfp_extracted else None

        bc_ext_via_rfp = rec_by_rfp["ext_doc_no"] if rec_by_rfp else None
        bc_rfp_via_invoice = rec_by_invoice["rfp_no"] if rec_by_invoice else None

        check1 = "x"
        if bc_rfp_via_invoice is not None and rfp_extracted is not None:
            check1 = "+" if bc_rfp_via_invoice == rfp_extracted else "x"
        elif bc_rfp_via_invoice is None and rfp_extracted is None:
            check1 = ""

        check2 = "x"
        if bc_ext_via_rfp is not None and invoice_no is not None:
            check2 = "+" if bc_ext_via_rfp == invoice_no else "x"
        elif bc_ext_via_rfp is None and invoice_no is None:
            check2 = ""

        found_any = rec_by_invoice is not None or rec_by_rfp is not None
        agree = (check1 in ("+", "")) and (check2 in ("+", "")) and found_any
        discrepancy = found_any and not agree

        if not found_any:
            match_status = "Not found"
            stats["not_found"] += 1
        elif discrepancy:
            match_status = "Discrepancy"
            stats["discrepancy"] += 1
        else:
            match_status = "Matched"
            stats["matched"] += 1

        best_rec = rec_by_invoice or rec_by_rfp
        remaining_amount = best_rec["remaining_amount"] if best_rec else None
        remaining_sales_amt = best_rec.get("remaining_sales_amt") if best_rec else None
        due_date = best_rec["due_date"] if best_rec else None

        if best_rec and (best_rec.get("vendor_name") or best_rec.get("vendor_no")):
            key = (best_rec.get("vendor_no"), best_rec.get("vendor_name"))
            vendor_counter[key] = vendor_counter.get(key, 0) + 1

        base_status = derive_payment_status(best_rec, today)
        if discrepancy:
            status_update = base_status + STATUS_DISCREPANCY_SUFFIX
        else:
            status_update = base_status

        rows_out.append(
            list(row)
            + [
                rfp_extracted,
                bc_ext_via_rfp,
                check1,
                bc_rfp_via_invoice,
                check2,
                match_status,
                remaining_amount,
                remaining_sales_amt,
                due_date,
                status_update,
            ]
        )

    output_path = _write_status_workbook(out_headers, rows_out, vendor_counter, output_dir)
    return stats, len(rows_out), output_path, len(vendor_counter)


# ---------------------------------------------------------------------------
# Shared Excel writer (used by process() and process_in_memory())
# ---------------------------------------------------------------------------

def _write_status_workbook(out_headers, rows_out, vendor_counter, output_dir):
    """Write a styled Bulk Status workbook and return the saved path."""
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Bulk Status"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    body_font = Font(name="Arial")
    status_fills = {
        "Paid": PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
        "Open – not yet due": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        "Overdue – payment pending": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
        "Not found in our records": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    }

    for c_idx, h in enumerate(out_headers, start=1):
        cell = out_ws.cell(row=1, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    out_ws.row_dimensions[1].height = 30

    status_col_idx = len(out_headers)       # "Status update" is always last
    due_date_col_idx = len(out_headers) - 1  # "BC: Due Date" is always second-to-last

    for r_idx, row_data in enumerate(rows_out, start=2):
        for c_idx, value in enumerate(row_data, start=1):
            cell = out_ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = body_font
            if c_idx == due_date_col_idx and isinstance(value, (datetime, date)):
                cell.number_format = "dd-mm-yyyy"
        status_value = row_data[-1]
        base_status = status_value.split(" (discrepancy")[0]
        fill = status_fills.get(base_status)
        if fill:
            out_ws.cell(row=r_idx, column=status_col_idx).fill = fill

    for c_idx, h in enumerate(out_headers, start=1):
        max_len = max(
            [len(str(h))]
            + [len(str(r[c_idx - 1])) for r in rows_out if r[c_idx - 1] is not None]
            + [8]
        )
        out_ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 2, 45)

    out_ws.freeze_panes = "A2"
    out_ws.auto_filter.ref = out_ws.dimensions

    vendor_no = vendor_name = None
    distinct_vendors = len(vendor_counter)
    if distinct_vendors == 1:
        (vendor_no, vendor_name) = next(iter(vendor_counter.keys()))

    os.makedirs(output_dir, exist_ok=True)
    filename = (
        "Status Overview - Multiple Venues.xlsx"
        if distinct_vendors > 1
        else build_output_filename(vendor_name, vendor_no)
    )
    output_path = os.path.join(output_dir, filename)
    out_wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# In-memory matching — no vendor Excel file needed (used for table + free-text)
# ---------------------------------------------------------------------------

_IN_MEMORY_INPUT_HEADERS = ["Invoice No.", "RFP No."]


def process_in_memory(pairs, by_ext, by_rfp, output_dir, today=None):
    """Match a list of {invoice_no, rfp_no} dicts against pre-loaded BC dicts.

    Returns (stats, out_headers, rows_out, output_path) where rows_out is a
    list of value lists aligned with out_headers — no need to re-read the Excel.
    """
    if today is None:
        today = date.today()

    out_headers = _IN_MEMORY_INPUT_HEADERS + [
        "RFP No. (extracted)",
        "BC: External Doc. No. (via RFP)",
        "Match check 1 (invoice ↔ RFP)",
        "BC: RFP No. (via invoice no.)",
        "Match check 2 (RFP ↔ invoice)",
        "Match status",
        "BC: Remaining Amount",
        "BC: Remaining Amt. Sales Invoice",
        "BC: Due Date",
        "Status update",
    ]

    rows_out = []
    stats = {"matched": 0, "discrepancy": 0, "not_found": 0}
    vendor_counter = {}

    for pair in pairs:
        invoice_no = norm(pair.get("invoice_no") or "")
        rfp_extracted = norm(pair.get("rfp_no") or "") or None

        rec_by_invoice = by_ext.get(invoice_no) if invoice_no else None
        rec_by_rfp = by_rfp.get(rfp_extracted) if rfp_extracted else None

        bc_ext_via_rfp = rec_by_rfp["ext_doc_no"] if rec_by_rfp else None
        bc_rfp_via_invoice = rec_by_invoice["rfp_no"] if rec_by_invoice else None

        # Cross-checks are only meaningful when both identifiers are provided.
        # If the vendor omitted one side (e.g. RFP-only input), leave that
        # check neutral ("") rather than flagging a false discrepancy ("x").
        check1 = ""
        if invoice_no and rfp_extracted:
            if bc_rfp_via_invoice is not None:
                check1 = "+" if bc_rfp_via_invoice == rfp_extracted else "x"

        check2 = ""
        if invoice_no and rfp_extracted:
            if bc_ext_via_rfp is not None:
                check2 = "+" if bc_ext_via_rfp == invoice_no else "x"

        found_any = rec_by_invoice is not None or rec_by_rfp is not None
        agree = (check1 in ("+", "")) and (check2 in ("+", "")) and found_any
        discrepancy = found_any and not agree

        if not found_any:
            match_status = "Not found"
            stats["not_found"] += 1
        elif discrepancy:
            match_status = "Discrepancy"
            stats["discrepancy"] += 1
        else:
            match_status = "Matched"
            stats["matched"] += 1

        best_rec = rec_by_invoice or rec_by_rfp
        remaining_amount = best_rec["remaining_amount"] if best_rec else None
        remaining_sales_amt = best_rec.get("remaining_sales_amt") if best_rec else None
        due_date = best_rec["due_date"] if best_rec else None

        if best_rec and (best_rec.get("vendor_name") or best_rec.get("vendor_no")):
            key = (best_rec.get("vendor_no"), best_rec.get("vendor_name"))
            vendor_counter[key] = vendor_counter.get(key, 0) + 1

        base_status = derive_payment_status(best_rec, today)
        status_update = base_status + STATUS_DISCREPANCY_SUFFIX if discrepancy else base_status

        rows_out.append([
            invoice_no or "",
            rfp_extracted or "",
            rfp_extracted,
            bc_ext_via_rfp,
            check1,
            bc_rfp_via_invoice,
            check2,
            match_status,
            remaining_amount,
            remaining_sales_amt,
            due_date,
            status_update,
        ])

    output_path = _write_status_workbook(out_headers, rows_out, vendor_counter, output_dir)
    return stats, out_headers, rows_out, output_path


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("bc_export", help="Path to the Business Central Vendor Ledger Entries export (.xlsx)")
    parser.add_argument("vendor_file", help="Path to the vendor/venue invoice+RFP overview (.xlsx)")
    parser.add_argument(
        "output_dir",
        help=(
            "Directory to write the resulting workbook into. The filename is always "
            "'Status Overview - <Vendor Name> - <Vendor No>.xlsx' — this is not "
            "configurable, per house naming convention."
        ),
    )
    parser.add_argument("--today", help="Override today's date (YYYY-MM-DD) for testing", default=None)
    args = parser.parse_args()

    output_dir = args.output_dir
    if os.path.splitext(output_dir)[1].lower() == ".xlsx":
        # Caller passed a filename instead of a directory — honour the folder,
        # override the filename, since the naming convention is mandatory.
        print(
            f"Note: '{output_dir}' looks like a filename, not a directory. "
            f"The output name is always 'Status Overview - <Vendor Name> - <Vendor No>.xlsx', "
            f"so writing into '{os.path.dirname(output_dir) or '.'}' instead."
        )
        output_dir = os.path.dirname(output_dir) or "."

    today = None
    if args.today:
        today = datetime.strptime(args.today, "%Y-%m-%d").date()

    stats, total, output_path, distinct_vendors = process(args.bc_export, args.vendor_file, output_dir, today=today)
    print(f"Processed {total} rows -> {output_path}")
    print(f"  Matched:     {stats['matched']}")
    print(f"  Discrepancy: {stats['discrepancy']}")
    print(f"  Not found:   {stats['not_found']}")
    if distinct_vendors > 1:
        print(
            f"  Note: matched rows span {distinct_vendors} different vendors — "
            f"named 'Status Overview - Multiple Venues.xlsx' rather than a single "
            f"vendor. If that wasn't intended, split the input file by vendor first."
        )
    elif distinct_vendors == 0:
        print("  Warning: no vendor could be identified from any matched row — filename uses placeholders.")


if __name__ == "__main__":
    sys.exit(main())
