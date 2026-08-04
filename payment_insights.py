"""
Payment Insights: aggregate views over the full Vendor Ledger export --
Due Today / Due in Next 7 Days / Overdue Invoices / Ready for Payment /
Customer Not Paid / Highest Payables by Vendor / Upcoming Cash Outflow --
plus six vendor-grouped additional reports. Everything here reads
already-loaded ledger rows (payment_category.load_ledger output) -- no AI
call anywhere in this module, and it runs over the WHOLE ledger, not just
whatever Payment Category happened to look up in a given session.

Due/Overdue scope is fixed to Rule 5: a row is "eligible" for these
categories ONLY when Remaining Amount Sales Invoice = 0 AND
Remaining Amt. (LCY) != 0 (see payment_category.is_ready_for_payment) --
never from a plain status column.
"""
from collections import defaultdict
from datetime import date, timedelta

import openpyxl

from payment_category import is_ready_for_payment, _to_number, _as_date

FILTERABLE_FIELDS = ["vendor", "invoice_no", "amount", "remaining_amt_lcy",
                      "remaining_sales_amt", "due_date", "status"]

ROW_DISPLAY_COLUMNS = [
    ("vendor", "Vendor"),
    ("invoice_no", "Invoice No."),
    ("rfp_no", "RFP No."),
    ("amount", "Amount"),
    ("date", "Date"),
    ("due_date", "Due Date"),
    ("status", "Status"),
    ("remaining_amt_lcy", "Remaining Amt. (LCY)"),
    ("remaining_sales_amt", "Remaining Amount Sales Invoice"),
]


def _display(value):
    if value is None or value == "":
        return ""
    return str(value)


def serialize_row(row):
    """Row dict -> plain JSON-safe dict of just the display columns (drops
    internal keys like "raw"/"due_date_raw"/"row_num" and any date objects)."""
    return {label: _display(row.get(key)) for key, label in ROW_DISPLAY_COLUMNS}


def filter_options(rows):
    """Distinct values per filterable field, for populating dropdowns --
    computed from the UNFILTERED row set so options never shrink out from
    under the user as they narrow their selection."""
    options = {}
    for field in FILTERABLE_FIELDS:
        seen_set = set()
        seen = []
        for row in rows:
            val = _display(row.get(field))
            if val == "" or val in seen_set:
                continue
            seen_set.add(val)
            seen.append(val)
        options[field] = sorted(seen)
    return options


def apply_filters(rows, filters):
    """filters: {field: [selected values]}. A field with an empty/missing
    selection is not filtered on. Multiple selected values within one field
    are OR'd; multiple fields are AND'd."""
    if not filters:
        return rows
    active = {f: set(v) for f, v in filters.items() if v}
    if not active:
        return rows
    out = []
    for row in rows:
        if all(_display(row.get(field)) in allowed for field, allowed in active.items()):
            out.append(row)
    return out


def _sum_lcy(rows):
    total = 0.0
    for r in rows:
        v = _to_number(r.get("remaining_amt_lcy"))
        if v:
            total += v
    return round(total, 2)


def _group_sum_by_vendor(rows):
    totals = defaultdict(float)
    for r in rows:
        v = _to_number(r.get("remaining_amt_lcy")) or 0
        totals[r.get("vendor") or "(unknown vendor)"] += v
    return sorted(totals.items(), key=lambda kv: kv[1], reverse=True)


def _group_count_by_vendor(rows):
    counts = defaultdict(int)
    for r in rows:
        counts[r.get("vendor") or "(unknown vendor)"] += 1
    return sorted(counts.items(), key=lambda kv: kv[1], reverse=True)


def _chart_from_pairs(pairs, limit=15):
    pairs = list(pairs)[:limit]
    return {"labels": [p[0] for p in pairs], "values": [round(p[1], 2) for p in pairs]}


CASH_OUTFLOW_BUCKET_ORDER = [
    "Overdue", "This week (0-7 days)", "Next 2 weeks (8-14 days)",
    "Next 30 days (15-30 days)", "Beyond 30 days", "No due date",
]


def compute_insights(rows, today=None):
    if today is None:
        today = date.today()

    def due(r):
        return _as_date(r.get("due_date_raw"))

    eligible = [r for r in rows if is_ready_for_payment(r.get("remaining_amt_lcy"), r.get("remaining_sales_amt"))]
    due_today = [r for r in eligible if due(r) == today]
    due_next_7 = [r for r in eligible if due(r) and today < due(r) <= today + timedelta(days=7)]
    overdue = [r for r in eligible if due(r) and due(r) < today]
    customer_not_paid = [r for r in rows if _to_number(r.get("remaining_sales_amt")) not in (None, 0)]

    highest_payables = _group_sum_by_vendor(eligible)

    buckets = {k: 0.0 for k in CASH_OUTFLOW_BUCKET_ORDER}
    for r in eligible:
        d = due(r)
        amt = _to_number(r.get("remaining_amt_lcy")) or 0
        if d is None:
            buckets["No due date"] += amt
            continue
        delta = (d - today).days
        if delta < 0:
            buckets["Overdue"] += amt
        elif delta <= 7:
            buckets["This week (0-7 days)"] += amt
        elif delta <= 14:
            buckets["Next 2 weeks (8-14 days)"] += amt
        elif delta <= 30:
            buckets["Next 30 days (15-30 days)"] += amt
        else:
            buckets["Beyond 30 days"] += amt
    cash_outflow_pairs = [(k, round(buckets[k], 2)) for k in CASH_OUTFLOW_BUCKET_ORDER]

    categories = {
        "due_today": {"label": "Due Today", "rows": due_today, "total_lcy": _sum_lcy(due_today),
                      "chart": _chart_from_pairs(_group_sum_by_vendor(due_today))},
        "due_next_7_days": {"label": "Due in Next 7 Days", "rows": due_next_7, "total_lcy": _sum_lcy(due_next_7),
                            "chart": _chart_from_pairs(_group_sum_by_vendor(due_next_7))},
        "overdue": {"label": "Overdue Invoices", "rows": overdue, "total_lcy": _sum_lcy(overdue),
                    "chart": _chart_from_pairs(_group_sum_by_vendor(overdue))},
        "ready_for_payment": {"label": "Ready for Payment", "rows": eligible, "total_lcy": _sum_lcy(eligible),
                              "chart": _chart_from_pairs(_group_sum_by_vendor(eligible))},
        "customer_not_paid": {"label": "Customer Not Paid", "rows": customer_not_paid,
                              "total_lcy": _sum_lcy(customer_not_paid),
                              "chart": _chart_from_pairs(_group_count_by_vendor(customer_not_paid))},
        "highest_payables_by_vendor": {"label": "Highest Payables by Vendor", "rows": eligible,
                                       "total_lcy": _sum_lcy(eligible),
                                       "chart": _chart_from_pairs(highest_payables)},
        "upcoming_cash_outflow": {"label": "Upcoming Cash Outflow", "rows": eligible,
                                  "total_lcy": _sum_lcy(eligible),
                                  "chart": _chart_from_pairs(cash_outflow_pairs, limit=len(cash_outflow_pairs))},
    }

    additional_reports = {
        "due_invoices_by_vendor": {"label": "Due invoices by vendor", "rows": due_next_7 + due_today},
        "overdue_invoices_by_vendor": {"label": "Overdue invoices by vendor", "rows": overdue},
        "top_payable_vendors": {"label": "Top payable vendors", "pairs": highest_payables},
        "customer_not_paid_invoices": {"label": "Customer-not-paid invoices", "rows": customer_not_paid},
        "ready_for_payment_invoices": {"label": "Ready-for-payment invoices", "rows": eligible},
        "upcoming_payment_schedule": {
            "label": "Upcoming payment schedule",
            "rows": sorted(eligible, key=lambda r: (due(r) is None, due(r) or today)),
        },
    }

    return categories, additional_reports


def write_insights_excel(categories, additional_reports, out_path):
    wb = openpyxl.Workbook()
    first = True
    for key, cat in categories.items():
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = cat["label"][:31]
        headers = [label for _, label in ROW_DISPLAY_COLUMNS]
        ws.append(headers)
        for row in cat["rows"]:
            ws.append([serialize_row(row).get(label, "") for label in headers])

    for key, rep in additional_reports.items():
        ws = wb.create_sheet(title=rep["label"][:31])
        if "pairs" in rep:
            ws.append(["Vendor", "Total Remaining Amt. (LCY)"])
            for vendor, total in rep["pairs"]:
                ws.append([vendor, total])
        else:
            headers = [label for _, label in ROW_DISPLAY_COLUMNS]
            ws.append(headers)
            for row in rep["rows"]:
                ws.append([serialize_row(row).get(label, "") for label in headers])

    wb.save(out_path)
    return out_path
