#!/usr/bin/env python3
"""
Local correction-detection engine for the AP dashboard.

Compares a *snapshot* of a bulk-status output workbook (saved at generation
time) against the *current* version of that same file on disk (which you may
have hand-edited in Excel), and produces structured correction records —
no LLM call anywhere in this script.

Matching key: invoice number (first column whose header contains
"invoice" — matches the vendor file's own header, whatever it was, since
bulk-status preserves original columns and appends its own after).

Tracked columns (the ones bulk-status actually appends — see
match_status.py):
    "Match status"   -> Matched / Discrepancy / Not found
    "Status update"  -> Paid / Open - not yet due / Overdue - payment pending /
                        Not found in our records  (+ optional discrepancy suffix)
Any other column edit is still captured, just tagged as a generic edit.

Usage:
    python detect_changes.py <snapshot.xlsx> <current.xlsx> <corrections_log.jsonl>

Appends one JSON object per changed cell to the log (JSONL, append-only).
Safe to run repeatedly — only genuinely new differences vs. the snapshot
are logged; run update_snapshot() afterwards if you want the current file
to become the new baseline (dashboard does this after you review a batch).
"""
import argparse
import json
import re
import sys
from datetime import datetime, date

import openpyxl

STATUS_COL_HINTS = ["status update"]
MATCH_COL_HINTS = ["match status"]
# "factuurnummer" is the Dutch vendor-file header bulk-status's own skill
# supports (see match_status.py's VENDOR_COLUMN_HINTS) -- without it, a
# Dutch-only invoice column is invisible to "invoice" and the search falls
# through to whatever appended column happens to also contain "invoice".
INVOICE_COL_HINTS = ["invoice", "factuurnummer"]
RFP_COL_HINTS = ["rfp no", "rfp number", "rfp id", "rfp no. (extracted)"]
# bulk-status appends columns like "Match check 1 (invoice <-> RFP)" that
# incidentally contain "invoice" too -- exclude known-generated-column
# markers so the search can't land on one of those instead of the original.
GENERATED_COLUMN_MARKERS = ["match check", "(via", "bc:"]


def find_col(headers, hints, exclude=None):
    exclude = exclude or []
    lower = [str(h).strip().lower() if h else "" for h in headers]
    for hint in hints:
        for i, h in enumerate(lower):
            if hint in h and not any(ex in h for ex in exclude):
                return i
    return None


def norm_value(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def load_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    inv_idx = find_col(headers, INVOICE_COL_HINTS, exclude=GENERATED_COLUMN_MARKERS)
    if inv_idx is None:
        raise ValueError(
            f"Could not find an 'invoice'-like column in {path}. "
            f"Headers found: {headers}"
        )
    rfp_idx = find_col(headers, RFP_COL_HINTS)
    rows_by_invoice = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        inv_key = norm_value(row[inv_idx])
        if inv_key is None:
            continue
        rows_by_invoice[inv_key] = {
            "headers": headers,
            "row": row,
            "rfp": norm_value(row[rfp_idx]) if rfp_idx is not None else None,
        }
    return headers, rows_by_invoice


def classify_correction(column_name, old_value, new_value):
    """Heuristic tagging — mechanical pattern match, no LLM."""
    col_lower = (column_name or "").lower()
    if "status update" in col_lower:
        # Category-shift is the interesting case for the classifier-learning loop
        old_base = (old_value or "").split(" (discrepancy")[0]
        new_base = (new_value or "").split(" (discrepancy")[0]
        if old_base != new_base:
            return "status_override"
        return "wording_edit"
    if "match status" in col_lower:
        return "match_correction"
    if "amount" in col_lower or "due date" in col_lower:
        return "amount_or_date_fix"
    return "generic_edit"


def diff_workbooks(snapshot_path, current_path):
    snap_headers, snap_rows = load_rows(snapshot_path)
    cur_headers, cur_rows = load_rows(current_path)

    corrections = []
    now = datetime.now().isoformat(timespec="seconds")

    for inv_key, cur_data in cur_rows.items():
        snap_data = snap_rows.get(inv_key)
        if snap_data is None:
            # New row not present in the original snapshot — not a "correction",
            # skip; it's a genuinely new invoice added after the fact.
            continue
        for col_idx, col_name in enumerate(cur_headers):
            if col_idx >= len(snap_data["row"]):
                continue
            old_val = norm_value(snap_data["row"][col_idx])
            new_val = norm_value(cur_data["row"][col_idx])
            if old_val != new_val:
                corrections.append({
                    "timestamp": now,
                    "snapshot_file": snapshot_path,
                    "current_file": current_path,
                    "invoice_no": inv_key,
                    "rfp_no": cur_data.get("rfp"),
                    "column": col_name,
                    "old_value": old_val,
                    "new_value": new_val,
                    "correction_type": classify_correction(col_name, old_val, new_val),
                    "reviewed": False,
                })
    return corrections


def append_log(corrections, log_path):
    with open(log_path, "a", encoding="utf-8") as f:
        for c in corrections:
            f.write(json.dumps(c, ensure_ascii=False) + "\n")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("snapshot", help="Path to the original snapshot workbook")
    parser.add_argument("current", help="Path to the current (possibly hand-edited) workbook")
    parser.add_argument("log", help="Path to corrections_log.jsonl to append to")
    args = parser.parse_args()

    corrections = diff_workbooks(args.snapshot, args.current)
    if not corrections:
        print("No changes detected.")
        return
    append_log(corrections, args.log)
    print(f"Detected {len(corrections)} cell-level change(s), appended to {args.log}")
    by_type = {}
    for c in corrections:
        by_type[c["correction_type"]] = by_type.get(c["correction_type"], 0) + 1
    for t, n in by_type.items():
        print(f"  {t}: {n}")


if __name__ == "__main__":
    sys.exit(main())
