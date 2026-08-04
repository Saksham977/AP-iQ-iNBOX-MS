#!/usr/bin/env python3
"""
Local AP dashboard server.

Run with:  python server.py
Then open: http://localhost:5000

Only two routes ever call the Claude API (both require you to click a button
in the UI — nothing here fires automatically):
    POST /api/update-classifier   -> batches unreviewed corrections, drafts a
                                      SKILL.md proposal via Claude, saved to
                                      data/proposals/ for your approval
    POST /api/draft-reply         -> drafts an email for one row via Claude

Everything else (running bulk-status, diffing Excel changes, listing
corrections, approving/rejecting proposals) is local Python — no API cost.
"""
import base64
import hashlib
import json
import mimetypes
import os
import pickle
import re
import subprocess
import sys
import glob
import urllib.error
import urllib.request
from datetime import datetime, date

import openpyxl
from flask import Flask, jsonify, request, send_from_directory

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
SNAPSHOTS_DIR = os.path.join(DATA_DIR, "snapshots")
PROPOSALS_DIR = os.path.join(DATA_DIR, "proposals")
CORRECTIONS_LOG = os.path.join(DATA_DIR, "corrections_log.jsonl")
PAYMENT_CATEGORY_DIR = os.path.join(DATA_DIR, "payment_category")
PAYMENT_CATEGORY_LOG = os.path.join(DATA_DIR, "payment_category_log.jsonl")
LEDGER_CACHE_DIR = os.path.join(DATA_DIR, "ledger_cache")
PAYMENT_INSIGHTS_DIR = os.path.join(DATA_DIR, "payment_insights")
LEARNED_RULES_FILE = os.path.join(DATA_DIR, "learned_rules.md")
STATIC_DIR = os.path.join(BASE_DIR, "static")

# Optional second-tier model provider -- any OpenAI-compatible chat
# completions endpoint (OpenAI itself, Azure OpenAI, Gemini's OpenAI-compat
# endpoint, OpenRouter, a local server like Ollama/LM Studio, etc). Only
# used when OTHER_MODEL_API_KEY is set; otherwise it's skipped entirely.
# See call_claude() / call_claude_extract() for where this sits in the
# fallback order: Claude API -> this provider -> local claude CLI.
OTHER_MODEL_API_KEY = os.environ.get("OTHER_MODEL_API_KEY")
OTHER_MODEL_BASE_URL = os.environ.get("OTHER_MODEL_BASE_URL", "https://api.openai.com/v1")
OTHER_MODEL_NAME = os.environ.get("OTHER_MODEL_NAME", "gpt-4o-mini")

SKILLS_MIRROR_DIR = os.path.join(BASE_DIR, "skills_mirror")
BULK_STATUS_SCRIPT = os.path.join(SKILLS_MIRROR_DIR, "bulk-status", "scripts", "match_status.py")
AP_CATEGORY_SKILL_MD = os.path.join(SKILLS_MIRROR_DIR, "ap-category-parameters", "SKILL.md")

sys.path.insert(0, os.path.join(BASE_DIR, "watcher"))
sys.path.insert(0, os.path.join(SKILLS_MIRROR_DIR, "bulk-status", "scripts"))
import detect_changes  # noqa: E402
import payment_category  # noqa: E402
import payment_insights  # noqa: E402
import match_status as bulk_matcher  # noqa: E402

app = Flask(__name__, static_folder=None)

# ---------------------------------------------------------------------------
# Unified ledger cache — one entry per (abs_path, mtime). A single file load
# produces both the payment-category LedgerIndex and the bulk-status dicts so
# the BC export is never read twice regardless of which panel calls first.
# ---------------------------------------------------------------------------
from collections import namedtuple as _namedtuple
_UnifiedLedger = _namedtuple("_UnifiedLedger", ["pc_index", "by_ext", "by_rfp"])
_UNIFIED_CACHE: dict = {}


def _load_unified_ledger(path):
    """Open the BC Vendor Ledger file once and build all views from a single parse."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    for sheet in wb.worksheets:
        if "vendor ledger" in sheet.title.lower():
            ws = sheet
            break
    header_row = bulk_matcher.find_header_row(ws, bulk_matcher.BC_COLUMN_HINTS, max_scan=1)
    all_rows = list(ws.iter_rows(min_row=header_row, values_only=True))
    wb.close()
    if not all_rows:
        empty_pc = payment_category.LedgerIndex(rows=[], by_invoice={}, by_rfp={})
        return _UnifiedLedger(pc_index=empty_pc, by_ext={}, by_rfp={})
    headers, data_rows = all_rows[0], all_rows[1:]
    pc_index = payment_category.load_ledger_from_rows(headers, data_rows)
    try:
        by_ext, by_rfp = bulk_matcher.build_bc_dicts_from_rows(headers, data_rows)
    except ValueError:
        # File lacks BC Vendor Ledger Entries columns (e.g. a vendor-format file).
        # pc_index is still valid; bulk-status lookups will return empty.
        by_ext, by_rfp = {}, {}
    return _UnifiedLedger(pc_index=pc_index, by_ext=by_ext, by_rfp=by_rfp)


def _ledger_pkl_path(abs_path, mtime):
    """Return the pickle path for a given source file + mtime.
    The path hash keeps filenames unique even when two different ledger files
    share the same basename (e.g. copies in different directories)."""
    path_hash = hashlib.md5(abs_path.encode("utf-8")).hexdigest()[:8]
    fname = os.path.splitext(os.path.basename(abs_path))[0]
    # Use integer milliseconds so the key is filesystem-safe and round-trips
    mtime_ms = int(mtime * 1000)
    return os.path.join(LEDGER_CACHE_DIR, f"{fname}_{path_hash}_{mtime_ms}.pkl")


def _get_unified_ledger(path):
    abs_path = os.path.abspath(path)
    mtime = os.path.getmtime(path)
    key = (abs_path, mtime)

    # 1. In-memory cache — fastest, valid for the current server session.
    if key in _UNIFIED_CACHE:
        return _UNIFIED_CACHE[key]

    os.makedirs(LEDGER_CACHE_DIR, exist_ok=True)
    pkl_path = _ledger_pkl_path(abs_path, mtime)

    # 2. Disk pickle cache — survives server restarts; loads in ~1–2 seconds
    #    even for a 50 MB+ ledger that takes minutes to parse from Excel.
    if os.path.exists(pkl_path):
        try:
            with open(pkl_path, "rb") as fh:
                ledger = pickle.load(fh)
            _UNIFIED_CACHE.clear()
            _UNIFIED_CACHE[key] = ledger
            print(f"[ledger] loaded from disk cache ({os.path.basename(pkl_path)})", flush=True)
            return ledger
        except Exception as exc:
            print(f"[ledger] disk cache corrupt, re-parsing ({exc})", flush=True)
            try:
                os.remove(pkl_path)
            except OSError:
                pass

    # 3. Parse from Excel — slow on first load (minutes for large files).
    #    After this completes the result is saved to disk so future starts
    #    (including after the server is restarted) use path 2 instead.
    size_mb = os.path.getsize(abs_path) / (1024 * 1024)
    print(f"[ledger] parsing {os.path.basename(abs_path)} ({size_mb:.1f} MB) — "
          "this may take a few minutes on the first load...", flush=True)
    _UNIFIED_CACHE.clear()
    ledger = _load_unified_ledger(abs_path)
    _UNIFIED_CACHE[key] = ledger

    # Persist to disk; clean up any stale pickles for this same source file.
    try:
        path_hash = hashlib.md5(abs_path.encode("utf-8")).hexdigest()[:8]
        fname_base = os.path.splitext(os.path.basename(abs_path))[0]
        pattern = os.path.join(LEDGER_CACHE_DIR, f"{fname_base}_{path_hash}_*.pkl")
        for old in glob.glob(pattern):
            if old != pkl_path:
                try:
                    os.remove(old)
                except OSError:
                    pass
        with open(pkl_path, "wb") as fh:
            pickle.dump(ledger, fh, protocol=pickle.HIGHEST_PROTOCOL)
        print(f"[ledger] cached to disk ({os.path.basename(pkl_path)})", flush=True)
    except Exception as exc:
        # Non-fatal: the in-memory cache still works for this session.
        print(f"[ledger] warning: could not save disk cache: {exc}", flush=True)

    return ledger


@app.route("/api/ledger-cache-status", methods=["GET"])
def ledger_cache_status():
    """
    Return whether the disk cache exists for the given ledger file so the UI
    can warn the user before the first (slow) parse.

    Query params: path=<abs_path_to_ledger_file>
    Response:
      { "cached": true/false,
        "in_memory": true/false,
        "size_mb": <float>,
        "cache_file": <basename or null> }
    """
    path = request.args.get("path", "").strip()
    if not path or not os.path.exists(path):
        return jsonify({"cached": False, "in_memory": False, "size_mb": 0, "cache_file": None})

    abs_path = os.path.abspath(path)
    try:
        mtime = os.path.getmtime(abs_path)
    except OSError:
        return jsonify({"cached": False, "in_memory": False, "size_mb": 0, "cache_file": None})

    key = (abs_path, mtime)
    in_memory = key in _UNIFIED_CACHE

    pkl_path = _ledger_pkl_path(abs_path, mtime)
    cached_on_disk = os.path.exists(pkl_path)

    size_mb = 0.0
    try:
        size_mb = round(os.path.getsize(abs_path) / (1024 * 1024), 1)
    except OSError:
        pass

    return jsonify({
        "cached": cached_on_disk or in_memory,
        "in_memory": in_memory,
        "size_mb": size_mb,
        "cache_file": os.path.basename(pkl_path) if cached_on_disk else None,
    })


@app.route("/api/clear-ledger-cache", methods=["POST"])
def clear_ledger_cache():
    """Delete all on-disk pickle files from LEDGER_CACHE_DIR and drop the
    in-memory cache.  The next request that needs the ledger will re-parse
    from Excel (slow first load) and rebuild the cache automatically."""
    _UNIFIED_CACHE.clear()

    removed = []
    errors = []
    if os.path.isdir(LEDGER_CACHE_DIR):
        for pkl in glob.glob(os.path.join(LEDGER_CACHE_DIR, "*.pkl")):
            try:
                os.remove(pkl)
                removed.append(os.path.basename(pkl))
            except OSError as exc:
                errors.append(f"{os.path.basename(pkl)}: {exc}")

    return jsonify({
        "removed": removed,
        "errors": errors,
        "message": (
            f"Cleared {len(removed)} cache file(s)."
            if not errors else
            f"Cleared {len(removed)} file(s); {len(errors)} could not be removed."
        ),
    })


@app.route("/api/claude-health", methods=["GET"])
def claude_health():
    """
    Lightweight check of which Claude tier is reachable — no LLM call made.
    Response: { tier, status, message }
      tier:   "api" | "other" | "cli" | "none"
      status: "connected" | "not_installed" | "not_authenticated" | "error"
    """
    import shutil as _shutil
    if os.environ.get("ANTHROPIC_API_KEY"):
        return jsonify({"tier": "api", "status": "connected",
                        "message": "Anthropic API key configured"})
    if OTHER_MODEL_API_KEY:
        return jsonify({"tier": "other", "status": "connected",
                        "message": f"Connected via {OTHER_MODEL_NAME}"})
    if not _shutil.which("claude"):
        return jsonify({"tier": "none", "status": "not_installed",
                        "message": (
                            "Claude CLI not found. Install it from claude.ai/code "
                            "and run 'claude' once to log in — or set ANTHROPIC_API_KEY."
                        )})
    try:
        r = subprocess.run(["claude", "--version"],
                           capture_output=True, text=True, timeout=8)
        if r.returncode == 0:
            return jsonify({"tier": "cli", "status": "connected",
                            "message": "Claude CLI connected (no API key needed)"})
        return jsonify({"tier": "cli", "status": "not_authenticated",
                        "message": (
                            "Claude CLI is installed but not logged in. "
                            "Run 'claude' in a terminal to authenticate."
                        )})
    except Exception as exc:
        return jsonify({"tier": "cli", "status": "error",
                        "message": f"Claude CLI error: {exc}"})


CLAUDE_MODEL = "claude-haiku-4-5-20251001"

_CLI_NOT_INSTALLED = (
    "Claude CLI is not installed. "
    "Install it from claude.ai/code and run 'claude' once to log in — "
    "or set the ANTHROPIC_API_KEY environment variable."
)
_CLI_NOT_AUTHENTICATED = (
    "Claude CLI is installed but not authenticated. "
    "Run 'claude' in a terminal to log in."
)


def _check_cli_result(result):
    """Raise a human-readable RuntimeError when the claude CLI subprocess fails."""
    if result.returncode != 0:
        stderr = result.stderr.strip()
        if "not logged in" in stderr.lower() or "authentication" in stderr.lower():
            raise RuntimeError(_CLI_NOT_AUTHENTICATED)
        raise RuntimeError(f"claude CLI failed: {stderr}")


def _other_provider_chat(messages, max_tokens, json_mode=False):
    """
    POST to an OpenAI-compatible /chat/completions endpoint. Works with
    OpenAI itself, Azure OpenAI, Gemini's OpenAI-compat endpoint,
    OpenRouter, or a local server (Ollama, LM Studio) -- anything that
    speaks this same request/response shape.
    """
    body = {"model": OTHER_MODEL_NAME, "max_tokens": max_tokens, "messages": messages}
    if json_mode:
        body["response_format"] = {"type": "json_object"}
    req = urllib.request.Request(
        f"{OTHER_MODEL_BASE_URL.rstrip('/')}/chat/completions",
        data=json.dumps(body).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {OTHER_MODEL_API_KEY}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", "ignore")[:300]
        raise RuntimeError(f"{OTHER_MODEL_NAME} call failed: {e.code} {detail}")
    except urllib.error.URLError as e:
        raise RuntimeError(f"{OTHER_MODEL_NAME} call failed: {e.reason}")
    return payload["choices"][0]["message"]["content"]


def call_claude(prompt, max_tokens=1500):
    """
    Get a text completion. Three-tier fallback:
      1. ANTHROPIC_API_KEY set -> Claude API directly.
      2. Else OTHER_MODEL_API_KEY set -> that OpenAI-compatible provider.
      3. Else -> the locally installed Claude Code CLI (`claude -p`), which
         uses your logged-in session and needs no API key at all -- this is
         unchanged from how the app has always run with no keys configured.
    Tiers 1 and 3 use Haiku at low effort -- these tasks (drafting a short
    email, proposing a detection block) don't need a bigger model, and it
    keeps token usage down. Tier 2 uses whatever OTHER_MODEL_NAME is set to.
    """
    if os.environ.get("ANTHROPIC_API_KEY"):
        import anthropic
        client = anthropic.Anthropic()
        response = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=max_tokens,
            messages=[{"role": "user", "content": prompt}],
        )
        return "".join(block.text for block in response.content if block.type == "text")

    if OTHER_MODEL_API_KEY:
        return _other_provider_chat([{"role": "user", "content": prompt}], max_tokens)

    try:
        result = subprocess.run(
            ["claude", "-p", "--model", CLAUDE_MODEL, "--effort", "low", "--setting-sources", "user"],
            input=prompt, capture_output=True, text=True, encoding="utf-8", timeout=120,
        )
    except FileNotFoundError:
        raise RuntimeError(_CLI_NOT_INSTALLED)
    _check_cli_result(result)
    return result.stdout.strip()


# ---------------------------------------------------------------------------
# Payment Category extraction: a SEPARATE model/helper from call_claude()
# above. Screenshot/Image and Email/Text extraction use Sonnet at low effort
# (per requirement); the existing classifier/draft-reply calls above are
# untouched and stay on Haiku.
# ---------------------------------------------------------------------------

CLAUDE_SONNET_MODEL = "claude-sonnet-4-6"

EXTRACTION_SCHEMA = {
    "type": "object",
    "properties": {
        "vendor": {"type": ["string", "null"]},
        "invoice_no": {"type": ["string", "null"]},
        "rfp_no": {"type": ["string", "null"]},
        "amount": {"type": ["number", "null"]},
        "date": {"type": ["string", "null"]},
    },
    "required": ["vendor", "invoice_no", "rfp_no", "amount", "date"],
}

EXTRACTION_INSTRUCTIONS = (
    "Extract these AP invoice/payment fields if present, else use null for "
    "any field you cannot find -- do not guess or invent a value: "
    "vendor (company/supplier name), invoice_no (invoice number), "
    "rfp_no (RFP/purchase reference number, if any), amount (numeric total, "
    "no currency symbol), date (invoice or due date). "
    "Respond with ONLY a JSON object with exactly these keys: "
    "vendor, invoice_no, rfp_no, amount, date."
)

# Multi-invoice variant: used when free-form text may contain several invoices.
EXTRACTION_SCHEMA_MANY = {
    "type": "object",
    "properties": {
        "invoices": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "vendor": {"type": ["string", "null"]},
                    "invoice_no": {"type": ["string", "null"]},
                    "rfp_no": {"type": ["string", "null"]},
                    "amount": {"type": ["number", "null"]},
                    "date": {"type": ["string", "null"]},
                },
                "required": ["vendor", "invoice_no", "rfp_no", "amount", "date"],
            },
        }
    },
    "required": ["invoices"],
}

EXTRACTION_INSTRUCTIONS_MANY = (
    "The text below may contain one or more AP invoices. "
    "Extract EVERY invoice as a list — do not stop after the first one. "
    "For each invoice extract these fields if present, else use null: "
    "vendor (company/supplier name), "
    "invoice_no (copy the full invoice number exactly as written, do not truncate), "
    "rfp_no (RFP/purchase reference number, if any), "
    "amount (numeric total, no currency symbol — treat comma as decimal separator "
    "if needed, e.g. €301,20 → 301.20), "
    "date (invoice date, ISO format YYYY-MM-DD where possible). "
    "Respond with ONLY a JSON object with key 'invoices' whose value is an array "
    "of objects each with exactly these keys: vendor, invoice_no, rfp_no, amount, date."
)


def _extraction_instructions():
    """
    EXTRACTION_INSTRUCTIONS plus any rules learned from approved Payment
    Category corrections (see "Update AI Learning", decide_proposal above).
    Re-read on every call so newly approved rules apply immediately.
    """
    instructions = EXTRACTION_INSTRUCTIONS
    if os.path.exists(LEARNED_RULES_FILE):
        with open(LEARNED_RULES_FILE, encoding="utf-8") as f:
            learned = f.read().strip()
        if learned:
            instructions += (
                "\n\nRULES LEARNED FROM PAST CORRECTIONS -- apply these when relevant:\n" + learned
            )
    return instructions


def call_claude_extract(text=None, image_path=None):
    """
    Extract structured invoice/payment fields from either pasted text or a
    local image file, using Sonnet at low effort. Returns a dict with keys
    vendor/invoice_no/rfp_no/amount/date (any may be None).
    """
    instructions = _extraction_instructions()

    if os.environ.get("ANTHROPIC_API_KEY"):
        import anthropic
        client = anthropic.Anthropic()
        content = []
        if image_path:
            with open(image_path, "rb") as f:
                image_bytes = f.read()
            media_type = mimetypes.guess_type(image_path)[0] or "image/png"
            content.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": media_type,
                    "data": base64.b64encode(image_bytes).decode("ascii"),
                },
            })
        content.append({
            "type": "text",
            "text": instructions + (f"\n\nTEXT:\n{text}" if text else ""),
        })
        response = client.messages.create(
            model=CLAUDE_SONNET_MODEL,
            max_tokens=400,
            messages=[{"role": "user", "content": content}],
        )
        raw = "".join(block.text for block in response.content if block.type == "text")
        match = re.search(r"\{.*\}", raw, re.DOTALL)
        if not match:
            raise RuntimeError(f"Claude did not return JSON: {raw[:300]}")
        return json.loads(match.group(0))

    if OTHER_MODEL_API_KEY:
        content = [{"type": "text", "text": instructions + (f"\n\nTEXT:\n{text}" if text else "")}]
        if image_path:
            with open(image_path, "rb") as f:
                image_bytes = f.read()
            media_type = mimetypes.guess_type(image_path)[0] or "image/png"
            data_url = f"data:{media_type};base64,{base64.b64encode(image_bytes).decode('ascii')}"
            content.append({"type": "image_url", "image_url": {"url": data_url}})
        raw = _other_provider_chat([{"role": "user", "content": content}], max_tokens=400, json_mode=True)
        match = re.search(r"\{.*\}", raw, re.DOTALL)
        if not match:
            raise RuntimeError(f"{OTHER_MODEL_NAME} did not return JSON: {raw[:300]}")
        return json.loads(match.group(0))

    prompt = instructions
    args = ["claude", "-p", "--model", CLAUDE_SONNET_MODEL, "--effort", "low",
            "--output-format", "json", "--json-schema", json.dumps(EXTRACTION_SCHEMA),
            "--setting-sources", "user"]
    if image_path:
        prompt += f"\n\nRead the image at this local path and extract the fields from it: {image_path}"
        # bypassPermissions is enough for the Read tool to fire headlessly --
        # --allowedTools is a variadic flag and will swallow the prompt
        # argument that follows it if combined this way, so it's deliberately
        # omitted here.
        args += ["--permission-mode", "bypassPermissions"]
    else:
        prompt += f"\n\nTEXT:\n{text}"
    args.append(prompt)

    try:
        result = subprocess.run(args, capture_output=True, text=True, encoding="utf-8", timeout=120)
    except FileNotFoundError:
        raise RuntimeError(_CLI_NOT_INSTALLED)
    _check_cli_result(result)
    try:
        payload = json.loads(result.stdout)
        return payload["structured_output"]
    except (json.JSONDecodeError, KeyError) as e:
        raise RuntimeError(f"claude CLI returned unexpected output: {e}")


def _extraction_instructions_many():
    instructions = EXTRACTION_INSTRUCTIONS_MANY
    if os.path.exists(LEARNED_RULES_FILE):
        with open(LEARNED_RULES_FILE, encoding="utf-8") as f:
            learned = f.read().strip()
        if learned:
            instructions += (
                "\n\nRULES LEARNED FROM PAST CORRECTIONS -- apply these when relevant:\n" + learned
            )
    return instructions


def call_claude_extract_many(text):
    """
    Extract ALL invoices from free-form text. Returns a list of dicts
    (each with vendor/invoice_no/rfp_no/amount/date keys).
    """
    instructions = _extraction_instructions_many()
    full_prompt = instructions + f"\n\nTEXT:\n{text}"

    if os.environ.get("ANTHROPIC_API_KEY"):
        import anthropic
        client = anthropic.Anthropic()
        response = client.messages.create(
            model=CLAUDE_SONNET_MODEL,
            max_tokens=2000,
            messages=[{"role": "user", "content": full_prompt}],
        )
        raw = "".join(block.text for block in response.content if block.type == "text")
        m = re.search(r"\{.*\}", raw, re.DOTALL)
        if not m:
            raise RuntimeError(f"Claude did not return JSON: {raw[:300]}")
        return json.loads(m.group(0)).get("invoices", [])

    if OTHER_MODEL_API_KEY:
        raw = _other_provider_chat(
            [{"role": "user", "content": [{"type": "text", "text": full_prompt}]}],
            max_tokens=2000, json_mode=True,
        )
        m = re.search(r"\{.*\}", raw, re.DOTALL)
        if not m:
            raise RuntimeError(f"{OTHER_MODEL_NAME} did not return JSON: {raw[:300]}")
        return json.loads(m.group(0)).get("invoices", [])

    args = ["claude", "-p", "--model", CLAUDE_SONNET_MODEL, "--effort", "low",
            "--output-format", "json", "--json-schema", json.dumps(EXTRACTION_SCHEMA_MANY),
            "--setting-sources", "user", full_prompt]
    try:
        result = subprocess.run(args, capture_output=True, text=True, encoding="utf-8", timeout=120)
    except FileNotFoundError:
        raise RuntimeError(_CLI_NOT_INSTALLED)
    _check_cli_result(result)
    try:
        payload = json.loads(result.stdout)
        return payload["structured_output"].get("invoices", [])
    except (json.JSONDecodeError, KeyError) as e:
        raise RuntimeError(f"claude CLI returned unexpected output: {e}")


# ---------------------------------------------------------------------------
# Static dashboard
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return send_from_directory(STATIC_DIR, "dashboard.html")


@app.route("/static/<path:path>")
def static_files(path):
    return send_from_directory(STATIC_DIR, path)


# ---------------------------------------------------------------------------
# Bulk-status: run the existing skill script locally, no LLM involved
# ---------------------------------------------------------------------------

def _read_excel_rows(path, max_rows=1000):
    """Read an output Excel workbook and return (headers, rows_as_list_of_dicts).
    All values are converted to strings for JSON serialisation."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(max_row=max_rows + 1, values_only=True))
    wb.close()
    if not all_rows:
        return [], []
    headers = [str(h) if h is not None else "" for h in all_rows[0]]
    records = []
    for row in all_rows[1:]:
        rec = {}
        for i, v in enumerate(row):
            if i >= len(headers):
                break
            if isinstance(v, datetime):
                rec[headers[i]] = v.strftime("%d-%m-%Y")
            elif isinstance(v, date):
                rec[headers[i]] = v.strftime("%d-%m-%Y")
            elif v is None:
                rec[headers[i]] = ""
            else:
                rec[headers[i]] = str(v)
        records.append(rec)
    return headers, records


# ---------------------------------------------------------------------------
# Bulk Status free-text extraction via Claude Sonnet 4.6 at low effort
# ---------------------------------------------------------------------------

BULK_INVOICE_EXTRACT_SCHEMA = {
    "type": "object",
    "properties": {
        "rows": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "invoice_no": {"type": ["string", "null"]},
                    "rfp_no": {"type": ["string", "null"]},
                },
                "required": ["invoice_no", "rfp_no"],
            },
        }
    },
    "required": ["rows"],
}

BULK_INVOICE_EXTRACT_INSTRUCTIONS = (
    "Extract EVERY invoice number (factuur number) and RFP number from the text below. "
    "The text may be a table, a list, or free-form. "
    "invoice_no: the invoice or factuur number — copy it exactly as written, do not truncate. "
    "rfp_no: the RFP, PO, or purchase reference number if present, else null. "
    "Respond with ONLY a JSON object with key 'rows' whose value is an array of objects "
    "each with exactly these keys: invoice_no, rfp_no."
)


def call_claude_extract_bulk_invoices(text):
    """Extract invoice/RFP numbers from free-form text using Sonnet 4.6 at low effort."""
    prompt = BULK_INVOICE_EXTRACT_INSTRUCTIONS + f"\n\nTEXT:\n{text}"

    if os.environ.get("ANTHROPIC_API_KEY"):
        import anthropic
        client = anthropic.Anthropic()
        response = client.messages.create(
            model=CLAUDE_SONNET_MODEL,
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = "".join(block.text for block in response.content if block.type == "text")
        m = re.search(r"\{.*\}", raw, re.DOTALL)
        if not m:
            raise RuntimeError(f"Claude did not return JSON: {raw[:300]}")
        return json.loads(m.group(0)).get("rows", [])

    if OTHER_MODEL_API_KEY:
        raw = _other_provider_chat(
            [{"role": "user", "content": [{"type": "text", "text": prompt}]}],
            max_tokens=1000, json_mode=True,
        )
        m = re.search(r"\{.*\}", raw, re.DOTALL)
        if not m:
            raise RuntimeError(f"{OTHER_MODEL_NAME} did not return JSON: {raw[:300]}")
        return json.loads(m.group(0)).get("rows", [])

    args = ["claude", "-p", "--model", CLAUDE_SONNET_MODEL, "--effort", "low",
            "--output-format", "json", "--json-schema", json.dumps(BULK_INVOICE_EXTRACT_SCHEMA),
            "--setting-sources", "user", prompt]
    try:
        result = subprocess.run(args, capture_output=True, text=True, encoding="utf-8", timeout=60)
    except FileNotFoundError:
        raise RuntimeError(_CLI_NOT_INSTALLED)
    _check_cli_result(result)
    try:
        payload = json.loads(result.stdout)
        return payload["structured_output"].get("rows", [])
    except (json.JSONDecodeError, KeyError) as e:
        raise RuntimeError(f"claude CLI returned unexpected output: {e}")


_VENDOR_COLS_SCHEMA = {
    "type": "object",
    "properties": {
        "invoice_no":  {"type": ["integer", "null"]},
        "rfp_no":      {"type": ["integer", "null"]},
        "description": {"type": ["integer", "null"]},
    },
    "required": ["invoice_no", "rfp_no", "description"],
}

_VENDOR_COLS_PROMPT_TMPL = (
    "You are given the column headers (and a few sample values) from a vendor/venue invoice file.\n\n"
    "Headers (0-based index): {headers_json}\n\n"
    "Sample values from the first rows:\n{samples_json}\n\n"
    "Identify the 0-based column index for each of the following fields:\n"
    "  invoice_no  — the invoice number, document reference, or bill number\n"
    "  rfp_no      — the RFP, event ID, booking reference, or PO number (null if not present)\n"
    "  description — a free-text description or memo column (null if not present)\n\n"
    "Return ONLY a JSON object with these exact keys: "
    "{{\"invoice_no\": <int or null>, \"rfp_no\": <int or null>, \"description\": <int or null>}}"
)


def _ai_identify_vendor_columns(headers, sample_rows):
    """Ask Claude to identify invoice_no / rfp_no / description column indices.
    Returns a dict like {'invoice_no': 0, 'rfp_no': 3, 'description': None}, or
    None when no AI provider is available."""
    indexed = {i: str(h) for i, h in enumerate(headers) if h is not None}
    samples = [
        {i: str(row[i]) for i in range(len(row)) if i < len(headers) and row[i] is not None}
        for row in (sample_rows or [])[:5]
    ]
    prompt = _VENDOR_COLS_PROMPT_TMPL.format(
        headers_json=json.dumps(indexed),
        samples_json=json.dumps(samples, indent=2),
    )

    raw = None
    if os.environ.get("ANTHROPIC_API_KEY"):
        import anthropic
        client = anthropic.Anthropic()
        resp = client.messages.create(
            model=CLAUDE_SONNET_MODEL,
            max_tokens=256,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = "".join(b.text for b in resp.content if b.type == "text")
    elif OTHER_MODEL_API_KEY:
        raw = _other_provider_chat(
            [{"role": "user", "content": [{"type": "text", "text": prompt}]}],
            max_tokens=256, json_mode=True,
        )
    else:
        try:
            result = subprocess.run(
                ["claude", "-p", "--model", CLAUDE_SONNET_MODEL, "--effort", "low",
                 "--setting-sources", "user", prompt],
                capture_output=True, text=True, encoding="utf-8", timeout=30,
            )
            raw = result.stdout if result.returncode == 0 else None
        except Exception:
            raw = None

    if not raw:
        return None
    m = re.search(r"\{[^{}]+\}", raw, re.DOTALL)
    if not m:
        return None
    try:
        mapping = json.loads(m.group(0))
        return {k: (int(v) if v is not None else None) for k, v in mapping.items()}
    except Exception:
        return None


@app.route("/api/run-bulk-status", methods=["POST"])
def run_bulk_status():
    import shutil
    payload = request.json or {}
    bc_export = (payload.get("bc_export_path") or "").strip()
    vendor_file = (payload.get("vendor_file_path") or "").strip()
    free_text = (payload.get("free_text") or "").strip()

    if not bc_export or not os.path.exists(bc_export):
        return jsonify({"error": "bc_export_path is required and must exist"}), 400

    os.makedirs(SNAPSHOTS_DIR, exist_ok=True)

    table_rows = payload.get("table_rows")  # list of {invoice_no, rfp_no} from UI table

    # Pre-load unified ledger — shared across all three input modes
    try:
        ledger = _get_unified_ledger(bc_export)
    except Exception as e:
        return jsonify({"error": f"Could not read BC ledger: {e}"}), 500

    if free_text or table_rows is not None:
        # Resolve pairs: Claude extraction (free text) or direct UI table (no Claude)
        if free_text:
            try:
                pairs = call_claude_extract_bulk_invoices(free_text)
            except Exception as e:
                return jsonify({"error": f"Claude extraction failed: {e}"}), 500
            if not pairs:
                return jsonify({"error": "No invoice or RFP numbers could be extracted from the text"}), 400
        else:
            pairs = [
                {"invoice_no": (r.get("invoice_no") or "").strip(),
                 "rfp_no": (r.get("rfp_no") or "").strip()}
                for r in (table_rows or [])
            ]
            pairs = [p for p in pairs if p["invoice_no"] or p["rfp_no"]]
            if not pairs:
                return jsonify({"error": "No invoice or RFP numbers entered in the table"}), 400

        # Match entirely in memory — no temp files, no file re-reads
        try:
            stats, out_headers, rows_out_vals, output_path = bulk_matcher.process_in_memory(
                pairs, ledger.by_ext, ledger.by_rfp, SNAPSHOTS_DIR
            )
        except Exception as e:
            return jsonify({"error": str(e)}), 500

        total = len(rows_out_vals)
        stamped = os.path.join(
            SNAPSHOTS_DIR,
            f"{os.path.splitext(os.path.basename(output_path))[0]}__{datetime.now():%Y%m%d-%H%M%S}.xlsx",
        )
        shutil.copy2(output_path, stamped)
        stdout = (
            f"Processed {total} rows -> {output_path}\n"
            f"  Matched:     {stats['matched']}\n"
            f"  Discrepancy: {stats['discrepancy']}\n"
            f"  Not found:   {stats['not_found']}"
        )
        # Rows returned directly from process_in_memory — no Excel re-read needed
        rows = [dict(zip(out_headers, r)) for r in rows_out_vals]
        # Convert date objects to strings for JSON
        for rec in rows:
            for k, v in rec.items():
                if isinstance(v, (datetime, date)):
                    rec[k] = v.strftime("%d-%m-%Y")
                elif v is None:
                    rec[k] = ""
                elif not isinstance(v, str):
                    rec[k] = str(v)
        return jsonify({
            "stdout": stdout,
            "output_path": output_path,
            "snapshot_path": stamped,
            "excel_filename": os.path.basename(output_path),
            "headers": out_headers,
            "rows": rows,
        })

    # --- Vendor-file path: direct in-process call (no subprocess) ---
    if not vendor_file or not os.path.exists(vendor_file):
        return jsonify({"error": "Provide vendor_file_path, free_text, or table_rows"}), 400

    def _run_process(vcols_override=None):
        return bulk_matcher.process(
            bc_export, vendor_file, SNAPSHOTS_DIR,
            _by_ext=ledger.by_ext, _by_rfp=ledger.by_rfp,
            _vcols_override=vcols_override,
        )

    try:
        stats, total, output_path, _ = _run_process()
    except ValueError as e:
        error = str(e)
        # BC-file-swapped check
        if "Could not find a column for" in error and "ext_doc_no" in error:
            return jsonify({"error": error + (
                " — this usually means the Business Central export and the "
                "Vendor/venue file got swapped in the two path fields above "
                "(the BC export needs columns like 'External Document No.' / "
                "'RFP No.'; check which file is in which box)"
            )}), 500
        # Column not identified — try the AI fallback
        if "Cannot identify" in error or "Could not find a column" in error:
            try:
                v_headers, v_samples, _ = bulk_matcher.read_vendor_file_info(vendor_file)
                vcols_override = _ai_identify_vendor_columns(v_headers, v_samples)
            except Exception:
                vcols_override = None
            if vcols_override and vcols_override.get("invoice_no") is not None:
                try:
                    stats, total, output_path, _ = _run_process(vcols_override)
                except Exception as e2:
                    return jsonify({"error": str(e2)}), 500
            else:
                return jsonify({"error": error}), 500
        else:
            return jsonify({"error": error}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    stamped = os.path.join(
        SNAPSHOTS_DIR,
        f"{os.path.splitext(os.path.basename(output_path))[0]}__{datetime.now():%Y%m%d-%H%M%S}.xlsx",
    )
    shutil.copy2(output_path, stamped)
    stdout = (
        f"Processed {total} rows -> {output_path}\n"
        f"  Matched:     {stats['matched']}\n"
        f"  Discrepancy: {stats['discrepancy']}\n"
        f"  Not found:   {stats['not_found']}"
    )
    headers, rows = _read_excel_rows(output_path)
    return jsonify({
        "stdout": stdout,
        "output_path": output_path,
        "snapshot_path": stamped,
        "excel_filename": os.path.basename(output_path),
        "headers": headers,
        "rows": rows,
    })


# ---------------------------------------------------------------------------
# Corrections: local diff only, no LLM
@app.route("/api/bulk-status/download/<path:filename>")
def bulk_status_download(filename):
    return send_from_directory(SNAPSHOTS_DIR, filename, as_attachment=True)


# ---------------------------------------------------------------------------

@app.route("/api/detect-changes", methods=["POST"])
def detect_changes_route():
    payload = request.json or {}
    snapshot_path = payload.get("snapshot_path")
    current_path = payload.get("current_path")
    if not snapshot_path or not current_path:
        return jsonify({"error": "snapshot_path and current_path are required"}), 400

    corrections = detect_changes.diff_workbooks(snapshot_path, current_path)
    if corrections:
        detect_changes.append_log(corrections, CORRECTIONS_LOG)
    return jsonify({"detected": len(corrections), "corrections": corrections})


def _read_all_corrections():
    corrections = []
    if os.path.exists(CORRECTIONS_LOG):
        with open(CORRECTIONS_LOG, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                corrections.append(json.loads(line))
    return corrections


@app.route("/api/corrections", methods=["GET"])
def list_corrections():
    """List all logged corrections, optionally filtered to unreviewed only."""
    unreviewed_only = request.args.get("unreviewed_only", "false").lower() == "true"
    corrections = _read_all_corrections()
    if unreviewed_only:
        corrections = [c for c in corrections if not c.get("reviewed")]
    return jsonify(corrections)


@app.route("/api/corrections/export", methods=["GET"])
def export_corrections():
    """Download the full lifetime correction history (both excel-diff and
    Payment Category sources) as one Excel file -- available anytime."""
    corrections = _read_all_corrections()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Correction History"
    headers = ["Timestamp", "User", "Source", "Invoice No", "RFP No", "Field",
               "Original Value", "Corrected Value", "Correction Type", "Reviewed", "Source Document"]
    ws.append(headers)
    for c in corrections:
        ws.append([
            c.get("timestamp", ""),
            c.get("user", ""),
            c.get("source", "excel_diff"),
            c.get("invoice_no", ""),
            c.get("rfp_no", ""),
            c.get("column", ""),
            c.get("old_value", ""),
            c.get("new_value", ""),
            c.get("correction_type", ""),
            "Yes" if c.get("reviewed") else "No",
            c.get("source_document", c.get("current_file", "")),
        ])

    os.makedirs(DATA_DIR, exist_ok=True)
    out_path = os.path.join(DATA_DIR, f"correction-history-{datetime.now():%Y%m%d-%H%M%S-%f}.xlsx")
    wb.save(out_path)
    return send_from_directory(DATA_DIR, os.path.basename(out_path), as_attachment=True,
                                download_name="ap-inbox-correction-history.xlsx")


# ---------------------------------------------------------------------------
# Classifier update: ONE explicit Claude call per batch, proposal-only
# ---------------------------------------------------------------------------

@app.route("/api/update-classifier", methods=["POST"])
def update_classifier():
    """
    Batches unreviewed corrections and asks Claude to propose an updated
    ap-category-parameters detection block, in the exact house format.
    This is the ONLY place a correction-driven Claude call happens, and it
    only happens when you click the button — never automatically.

    NOTE: this endpoint calls the Anthropic API directly. Set ANTHROPIC_API_KEY
    in your environment before using it. It never writes to SKILL.md itself —
    it only writes a proposal file for you to review in the dashboard.
    """
    payload = request.json or {}
    category = payload.get("category", "Payment Status")

    corrections = []
    if os.path.exists(CORRECTIONS_LOG):
        with open(CORRECTIONS_LOG, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                rec = json.loads(line)
                if not rec.get("reviewed") and rec.get("correction_type") in (
                    "status_override", "match_correction", "amount_or_date_fix"
                ):
                    corrections.append(rec)

    if not corrections:
        return jsonify({"message": "No unreviewed corrections to process."})

    current_block = ""
    if os.path.exists(AP_CATEGORY_SKILL_MD):
        with open(AP_CATEGORY_SKILL_MD, encoding="utf-8") as f:
            current_block = f.read()

    payment_category_corrections = [c for c in corrections if c.get("source") == "payment_category"]

    prompt = f"""You are updating AP Inbox's automated detection/extraction logic
based on real corrections a human made. There are two kinds of corrections in
the batch below, distinguished by "source":

- source "excel_diff" (or missing): a bulk-status output workbook cell was
  hand-corrected. correction_type status_override = the Status update
  category changed; match_correction = the Match status changed.
- source "payment_category": a Payment Category extraction/match result
  (from a screenshot, email text, PDF, or Excel) was hand-corrected.

Produce a response with exactly two sections, in this order:

1. A line "SKILL_BLOCK:" followed by an updated detection block for the
   classifier category "{category}", in the exact house format used in
   /mnt/skills/user/ap-category-parameters/SKILL.md and its house-format
   reference (the "**\u2726 Je prompt:**" block plus the Dutch distilled note).
   Base this ONLY on the "excel_diff" corrections in the batch. Compare
   against the current signals below, keep what holds, and propose only
   genuinely new/adjusted signal patterns. If there are no excel_diff
   corrections in the batch, write "No changes proposed." for this section.

2. A line "LEARNED_RULES:" followed by a short bullet list of general,
   reusable rules for extracting invoice/payment fields (vendor, invoice
   number, RFP number, amount, date) from screenshots and pasted email/text,
   derived ONLY from the "payment_category" corrections in the batch. Each
   bullet should generalize beyond the single invoice it came from (e.g. a
   vendor-name spelling variant to recognize, a date format quirk, a field
   that's commonly confused with another) -- don't just restate the raw
   correction. If there are no payment_category corrections in the batch,
   write "No new rules." for this section.

CURRENT SKILL.md:
{current_block}

CORRECTIONS BATCH:
{json.dumps(corrections, indent=2, ensure_ascii=False)}
"""

    try:
        proposed_text = call_claude(prompt, max_tokens=1800)
    except Exception as e:
        return jsonify({"error": f"Claude call failed: {e}"}), 500

    skill_block = ""
    learned_rules_text = ""
    skill_match = re.search(r"SKILL_BLOCK:(.*?)(?=LEARNED_RULES:|$)", proposed_text, re.DOTALL)
    rules_match = re.search(r"LEARNED_RULES:(.*)$", proposed_text, re.DOTALL)
    if skill_match:
        skill_block = skill_match.group(1).strip()
    if rules_match:
        learned_rules_text = rules_match.group(1).strip()
    if not skill_match and not rules_match:
        # Claude didn't follow the section format -- keep the raw text visible rather than losing it.
        skill_block = proposed_text

    os.makedirs(PROPOSALS_DIR, exist_ok=True)
    proposal_path = os.path.join(
        PROPOSALS_DIR, f"{datetime.now():%Y-%m-%d-%H%M%S}-{category.lower().replace(' ', '-')}.json"
    )
    proposal = {
        "category": category,
        "created": datetime.now().isoformat(timespec="seconds"),
        "corrections_used": corrections,
        "payment_category_correction_count": len(payment_category_corrections),
        "proposed_block": skill_block,
        "learned_rules_text": learned_rules_text,
        "status": "pending",
    }
    with open(proposal_path, "w", encoding="utf-8") as f:
        json.dump(proposal, f, indent=2, ensure_ascii=False)

    return jsonify({"proposal_path": proposal_path, "proposal": proposal})


@app.route("/api/proposals", methods=["GET"])
def list_proposals():
    proposals = []
    for path in sorted(glob.glob(os.path.join(PROPOSALS_DIR, "*.json"))):
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        data["_path"] = path
        proposals.append(data)
    return jsonify(proposals)


@app.route("/api/proposals/<path:filename>/decide", methods=["POST"])
def decide_proposal(filename):
    """
    approve -> marks proposal approved (actual SKILL.md edit still needs to
               happen in a Claude session with str_replace access to
               /mnt/skills/user/ — this dashboard only tracks the decision
               and re-marks the source corrections as reviewed)
    reject  -> marks rejected, corrections marked reviewed so they aren't
               re-proposed identically next time
    """
    payload = request.json or {}
    decision = payload.get("decision")  # "approve" or "reject"
    if decision not in ("approve", "reject"):
        return jsonify({"error": "decision must be 'approve' or 'reject'"}), 400

    proposal_path = os.path.join(PROPOSALS_DIR, filename)
    if not os.path.exists(proposal_path):
        return jsonify({"error": "proposal not found"}), 404

    with open(proposal_path, encoding="utf-8") as f:
        proposal = json.load(f)
    proposal["status"] = "approved" if decision == "approve" else "rejected"
    proposal["decided_at"] = datetime.now().isoformat(timespec="seconds")
    with open(proposal_path, "w", encoding="utf-8") as f:
        json.dump(proposal, f, indent=2, ensure_ascii=False)

    learned_rules_text = (proposal.get("learned_rules_text") or "").strip()
    if decision == "approve" and learned_rules_text and learned_rules_text != "No new rules.":
        with open(LEARNED_RULES_FILE, "a", encoding="utf-8") as f:
            f.write(f"\n<!-- approved {proposal['decided_at']}, category: {proposal['category']} -->\n")
            f.write(learned_rules_text + "\n")

    # Mark the underlying corrections as reviewed either way, so a rejected
    # batch doesn't get re-proposed verbatim next time.
    used_invoices = {(c["invoice_no"], c["column"], c["timestamp"]) for c in proposal["corrections_used"]}
    if os.path.exists(CORRECTIONS_LOG):
        lines = []
        with open(CORRECTIONS_LOG, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                rec = json.loads(line)
                key = (rec.get("invoice_no"), rec.get("column"), rec.get("timestamp"))
                if key in used_invoices:
                    rec["reviewed"] = True
                lines.append(json.dumps(rec, ensure_ascii=False))
        with open(CORRECTIONS_LOG, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")

    return jsonify({"status": proposal["status"]})


# ---------------------------------------------------------------------------
# Email drafting: ONE explicit Claude call per row, on click only
# ---------------------------------------------------------------------------

@app.route("/api/draft-reply", methods=["POST"])
def draft_reply():
    """
    Drafts a reply email for a single row. Uses ONLY:
      - the row's current + prior status (from corrections log, if any)
      - optional vendor email text pasted in by the user for this reply

    Never assumes vendor email content that wasn't provided.
    """
    payload = request.json or {}
    invoice_no = payload.get("invoice_no")
    rfp_no = payload.get("rfp_no")
    current_status = payload.get("current_status")  # e.g. "Overdue - payment pending"
    match_status = payload.get("match_status")       # e.g. "Discrepancy"
    vendor_email_text = payload.get("vendor_email_text")  # optional, user-pasted

    if not invoice_no or not current_status:
        return jsonify({"error": "invoice_no and current_status are required"}), 400

    # Pull this invoice's correction history for context, if any
    history = []
    if os.path.exists(CORRECTIONS_LOG):
        with open(CORRECTIONS_LOG, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                rec = json.loads(line)
                if rec.get("invoice_no") == invoice_no:
                    history.append(rec)

    context_lines = [
        f"Invoice number: {invoice_no}",
        f"RFP number: {rfp_no or 'n/a'}",
        f"Current Status update: {current_status}",
        f"Current Match status: {match_status or 'n/a'}",
    ]
    if history:
        context_lines.append("Correction history for this invoice:")
        for h in history:
            context_lines.append(
                f"  - {h['column']} changed from '{h['old_value']}' to '{h['new_value']}' ({h['correction_type']})"
            )
    if vendor_email_text:
        context_lines.append("Vendor's email text for this specific reply (use only what's relevant, don't invent beyond it):")
        context_lines.append(vendor_email_text)
    else:
        context_lines.append(
            "No vendor email text was provided for this reply -- do not invent or assume "
            "what the vendor said. Draft generically off the row data only."
        )

    prompt = f"""Draft a short, professional AP reply email in English, in the
same register as these two house examples (do not copy them verbatim,
adapt tone and structure to what actually applies to this row):

EXAMPLE A (status overview, outbound):
"Dear sir/madam, Please find the attached overview of the invoices. Based on
our records, some of the invoices are not yet due for payment. Should you
have any questions or require any further clarification, please do not
hesitate to contact us. Hoping to have fully informed you."

EXAMPLE B (resolution/thank-you, after a discrepancy was resolved):
"Dear Sir/Madam, Thank you for your detailed explanation and for taking the
time to share the additional invoices. I can confirm that there is now no
discrepancy regarding this RFP. We will make every effort to process the
outstanding invoices as soon as possible. We greatly appreciate your
patience, and continued cooperation throughout this process. Thank you
again for your assistance."

Choose whichever pattern fits this row's actual data below, or blend as
needed -- e.g. if Match status shows a Discrepancy was corrected to Matched,
lean toward Example B's resolution tone; if this is a fresh status
communication, lean toward Example A but state the SPECIFIC status
(Paid / Overdue / Open-not-yet-due / Discrepancy) rather than the generic
"some invoices are not yet due" line. Be concrete about this row, not generic.

Output ONLY the email body -- no preamble, no notes or explanation of your
approach, no markdown formatting.

ROW CONTEXT:
{chr(10).join(context_lines)}
"""

    try:
        draft_text = call_claude(prompt, max_tokens=600)
    except Exception as e:
        return jsonify({"error": f"Claude call failed: {e}"}), 500

    return jsonify({"draft": draft_text})


# ---------------------------------------------------------------------------
# Payment Category: match a screenshot / pasted email-text / PDF / Excel
# invoice against the Vendor Ledger (the BC export from panel 1). Image and
# free-text go through Claude Sonnet at low effort (call_claude_extract,
# above); PDF and Excel are parsed and matched directly -- no Claude call.
# Everything stays on local file paths, nothing is uploaded to the server.
# ---------------------------------------------------------------------------

def _log_payment_category(entries):
    os.makedirs(PAYMENT_CATEGORY_DIR, exist_ok=True)
    with open(PAYMENT_CATEGORY_LOG, "a", encoding="utf-8") as f:
        for entry in entries:
            f.write(json.dumps(entry, ensure_ascii=False, default=str) + "\n")


@app.route("/api/payment-category/run", methods=["POST"])
def payment_category_run():
    payload = request.json or {}
    input_type = payload.get("input_type")
    ledger_path = payload.get("ledger_path")

    if input_type not in ("image", "text", "pdf", "excel"):
        return jsonify({"error": "input_type must be one of image, text, pdf, excel"}), 400
    if not ledger_path or not os.path.exists(ledger_path):
        return jsonify({"error": "ledger_path is required and must exist (the BC export / Vendor Ledger file)"}), 400

    try:
        ledger_index = _get_unified_ledger(ledger_path).pc_index
    except Exception as e:
        return jsonify({"error": f"Could not read ledger file: {e}"}), 500

    records = []
    try:
        if input_type == "image":
            image_path = payload.get("image_path")
            if not image_path or not os.path.exists(image_path):
                return jsonify({"error": "image_path is required and must exist"}), 400
            extracted = call_claude_extract(image_path=image_path)
            ledger_row = (payment_category.match_by_invoice_no(ledger_index, extracted.get("invoice_no"))
                          or payment_category.match_by_rfp_no(ledger_index, extracted.get("rfp_no")))
            records.append(payment_category.build_record("Image", extracted, ledger_row))

        elif input_type == "text":
            text_mode = payload.get("text_mode")
            text_value = (payload.get("text_value") or "").strip()
            if text_mode not in ("number", "text"):
                return jsonify({"error": "text_mode must be 'number' or 'text'"}), 400
            if not text_value:
                return jsonify({"error": "text_value is required"}), 400

            if text_mode == "number":
                # A bare invoice or RFP number -- no Claude call, direct lookup.
                extracted = {"vendor": None, "invoice_no": text_value, "rfp_no": text_value,
                             "amount": None, "date": None}
                ledger_row = (payment_category.match_by_invoice_no(ledger_index, text_value)
                              or payment_category.match_by_rfp_no(ledger_index, text_value))
                records.append(payment_category.build_record("Text", extracted, ledger_row))
            else:
                # Free-form text may contain many invoices -- extract all of them.
                extracted_list = call_claude_extract_many(text_value)
                if not extracted_list:
                    return jsonify({"error": "No invoices could be extracted from the text"}), 400
                for extracted in extracted_list:
                    ledger_row = (payment_category.match_by_invoice_no(ledger_index, extracted.get("invoice_no"))
                                  or payment_category.match_by_rfp_no(ledger_index, extracted.get("rfp_no")))
                    records.append(payment_category.build_record("Text", extracted, ledger_row))

        elif input_type in ("pdf", "excel"):
            file_path = payload.get("file_path")
            if not file_path or not os.path.exists(file_path):
                return jsonify({"error": "file_path is required and must exist"}), 400
            extracted_rows = (payment_category.extract_from_pdf(file_path) if input_type == "pdf"
                               else payment_category.extract_from_excel(file_path))
            if not extracted_rows:
                return jsonify({"error": "No invoice rows could be parsed from that file"}), 400
            source_label = "PDF" if input_type == "pdf" else "Excel"
            for row in extracted_rows:
                ledger_row = (payment_category.match_by_invoice_no(ledger_index, row.get("invoice_no"))
                              or payment_category.match_by_rfp_no(ledger_index, row.get("rfp_no")))
                records.append(payment_category.build_record(source_label, row, ledger_row))
    except Exception as e:
        return jsonify({"error": f"Extraction/matching failed: {e}"}), 500

    os.makedirs(PAYMENT_CATEGORY_DIR, exist_ok=True)
    filename = f"payment-category-{datetime.now():%Y%m%d-%H%M%S-%f}.xlsx"
    out_path = os.path.join(PAYMENT_CATEGORY_DIR, filename)
    payment_category.write_results_excel(records, out_path)

    warnings = [f"No ledger match found for invoice '{r['Invoice No.']}'" for r in records if r["Match"] == "No"]

    _log_payment_category([{
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "input_type": input_type,
        "ledger_path": ledger_path,
        "result": {k: v for k, v in r.items() if k != "_extra"},
    } for r in records])

    return jsonify({
        "records": [{k: v for k, v in r.items() if k != "_extra"} for r in records],
        "warnings": warnings,
        "excel_filename": filename,
        "output_path": out_path,
    })


@app.route("/api/payment-category/download/<filename>", methods=["GET"])
def payment_category_download(filename):
    return send_from_directory(PAYMENT_CATEGORY_DIR, filename, as_attachment=True)


# ---------------------------------------------------------------------------
# Payment Insights: aggregate views over the WHOLE ledger (not just whatever
# Payment Category looked up this session). No AI call -- pure local
# aggregation over payment_category.load_ledger() rows. See
# payment_insights.py for the category/report definitions.
# ---------------------------------------------------------------------------

@app.route("/api/payment-insights/run", methods=["POST"])
def payment_insights_run():
    payload = request.json or {}
    ledger_path = payload.get("ledger_path")
    filters = payload.get("filters") or {}

    if not ledger_path or not os.path.exists(ledger_path):
        return jsonify({"error": "ledger_path is required and must exist (the BC export / Vendor Ledger file)"}), 400

    try:
        all_rows = _get_unified_ledger(ledger_path).pc_index.rows
    except Exception as e:
        return jsonify({"error": f"Could not read ledger file: {e}"}), 500

    options = payment_insights.filter_options(all_rows)
    filtered_rows = payment_insights.apply_filters(all_rows, filters)
    categories, additional_reports = payment_insights.compute_insights(filtered_rows)

    os.makedirs(PAYMENT_INSIGHTS_DIR, exist_ok=True)
    filename = f"payment-insights-{datetime.now():%Y%m%d-%H%M%S-%f}.xlsx"
    out_path = os.path.join(PAYMENT_INSIGHTS_DIR, filename)
    payment_insights.write_insights_excel(categories, additional_reports, out_path)

    def cat_json(cat):
        return {
            "label": cat["label"],
            "total_lcy": cat["total_lcy"],
            "count": len(cat["rows"]),
            "rows": [payment_insights.serialize_row(r) for r in cat["rows"]],
            "chart": cat["chart"],
        }

    def report_json(rep):
        out = {"label": rep["label"]}
        if "pairs" in rep:
            out["pairs"] = [{"vendor": v, "total_lcy": t} for v, t in rep["pairs"]]
        else:
            out["rows"] = [payment_insights.serialize_row(r) for r in rep["rows"]]
        return out

    return jsonify({
        "filter_options": options,
        "categories": {k: cat_json(v) for k, v in categories.items()},
        "additional_reports": {k: report_json(v) for k, v in additional_reports.items()},
        "row_count": len(filtered_rows),
        "total_row_count": len(all_rows),
        "excel_filename": filename,
    })


@app.route("/api/payment-insights/download/<filename>", methods=["GET"])
def payment_insights_download(filename):
    return send_from_directory(PAYMENT_INSIGHTS_DIR, filename, as_attachment=True)


# ---------------------------------------------------------------------------
# Learning & Corrections: capturing a hand-fix to a Payment Category result.
# Appends to the SAME corrections_log.jsonl used by Excel-diff corrections
# (source="payment_category" distinguishes the two), so both feed the same
# "Update AI Learning" batch (see update_classifier above). No Claude call.
# ---------------------------------------------------------------------------

FIELD_TO_CORRECTION_TYPE = {
    "Status": "status_override",
    "Vendor": "match_correction",
    "Invoice No.": "match_correction",
    "RFP No.": "match_correction",
    "Ledger Ref.": "match_correction",
    "Amount": "amount_or_date_fix",
    "Date": "amount_or_date_fix",
    "Due Date": "amount_or_date_fix",
    "Remaining Amt. (LCY)": "amount_or_date_fix",
    "Remaining Amount Sales Invoice": "amount_or_date_fix",
}


@app.route("/api/payment-category/correct", methods=["POST"])
def payment_category_correct():
    payload = request.json or {}
    field = payload.get("field")
    new_value = payload.get("new_value")
    if not field or new_value is None or str(new_value).strip() == "":
        return jsonify({"error": "field and new_value are required"}), 400

    entry = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "source": "payment_category",
        "user": (payload.get("user") or "").strip(),
        "invoice_no": payload.get("invoice_no") or "",
        "rfp_no": payload.get("rfp_no") or "",
        "column": field,
        "old_value": payload.get("old_value", ""),
        "new_value": new_value,
        "correction_type": FIELD_TO_CORRECTION_TYPE.get(field, "generic_edit"),
        "source_document": payload.get("source_document") or "",
        "input_type": payload.get("input_type") or "",
        "reviewed": False,
    }
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(CORRECTIONS_LOG, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")

    return jsonify({"logged": entry})


if __name__ == "__main__":
    import threading
    import webbrowser

    os.makedirs(SNAPSHOTS_DIR, exist_ok=True)
    os.makedirs(PROPOSALS_DIR, exist_ok=True)
    os.makedirs(PAYMENT_CATEGORY_DIR, exist_ok=True)

    # Open the dashboard in the default browser ~1 s after Flask has started.
    def _open_browser():
        import time
        time.sleep(1.2)
        webbrowser.open("http://localhost:5000")

    threading.Thread(target=_open_browser, daemon=True).start()
    print("AP iQ iNBOX — opening http://localhost:5000 …", flush=True)
    app.run(debug=False, port=5000)
