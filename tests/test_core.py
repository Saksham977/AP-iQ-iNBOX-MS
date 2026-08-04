"""
Core unit tests for the AP Dashboard back-end.

Covers:
  1. match_status  — column detection, BC dict building, status derivation,
                     process_in_memory matching logic
  2. payment_category — payment-status rules, load_ledger_from_rows, matching
  3. server (unified cache) — _load_unified_ledger, _get_unified_ledger,
                              cache hit / miss / invalidation

All tests are purely in-memory; no real Excel files or network calls required.
"""
import io
import os
import sys
import shutil
import tempfile
import types
from collections import namedtuple
from datetime import date, datetime, timedelta
from unittest.mock import patch

import openpyxl
import pytest

# ---------------------------------------------------------------------------
# Path setup — make the project modules importable
# ---------------------------------------------------------------------------
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRIPTS_DIR = os.path.join(ROOT, "skills_mirror", "bulk-status", "scripts")
for p in (ROOT, SCRIPTS_DIR):
    if p not in sys.path:
        sys.path.insert(0, p)

import match_status as ms
import payment_category as pc

# ---------------------------------------------------------------------------
# Helpers — build tiny in-memory Excel files and raw-row fixtures
# ---------------------------------------------------------------------------

BC_HEADERS = (
    "External Document No.",
    "RFP No.",
    "Document No.",
    "Vendor No.",
    "Vendor Name",
    "Description",
    "Remaining Amount",
    "Remaining Amount Sales Invoice",
    "Due Date",
    "Open",
    "Currency Code",
)

VENDOR_HEADERS = ("Invoice No.", "RFP No.")


def _make_bc_rows(*data_rows):
    """Return (headers_tuple, list_of_data_tuples)."""
    return BC_HEADERS, list(data_rows)


def _make_wb_bytes(sheet_data):
    """Write sheet_data (list of row-tuples) to an in-memory xlsx and return bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in sheet_data:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _save_wb(sheet_data, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in sheet_data:
        ws.append(list(row))
    wb.save(path)


TODAY = date(2026, 8, 3)
PAST = TODAY - timedelta(days=5)
FUTURE = TODAY + timedelta(days=10)

# A single realistic BC ledger row
BC_ROW_OPEN_FUTURE = (
    "INV-001", "RFP-100", "P001", "V001", "Hotel Alpha", "Invoice",
    -500.0, 600.0, FUTURE, True, "",
)
BC_ROW_OPEN_PAST = (
    "INV-002", "RFP-200", "P002", "V001", "Hotel Alpha", "Invoice",
    -300.0, 0.0, PAST, True, "",
)
BC_ROW_PAID = (
    "INV-003", "RFP-300", "P003", "V002", "Hotel Beta", "Invoice",
    0.0, 0.0, PAST, False, "",
)
BC_ROW_NO_SALES = (
    "INV-004", "RFP-400", "P004", "V002", "Hotel Beta", "Invoice",
    -200.0, None, FUTURE, True, "",
)


# ===========================================================================
# 1. match_status — column detection
# ===========================================================================

class TestFindColumn:
    def test_exact_hint_match(self):
        headers = ["External Document No.", "RFP No.", "Vendor Name"]
        assert ms.find_column(headers, ["external document no"]) == 0

    def test_substring_hint_match(self):
        headers = ["Some Col", "RFP Number X", "Amount"]
        assert ms.find_column(headers, ["rfp no", "rfp number"]) == 1

    def test_exclude_prevents_match(self):
        headers = ["Remaining Amount Sales Invoice", "Remaining Amount"]
        # "remaining amount" hint should skip the sales-invoice column
        idx = ms.find_column(headers, ["remaining amount"], exclude=["sales invoice"])
        assert idx == 1

    def test_not_found_required_raises(self):
        with pytest.raises(ValueError, match="Could not find"):
            ms.find_column(["Col A", "Col B"], ["missing hint"], required=True, label="test")

    def test_not_found_optional_returns_none(self):
        assert ms.find_column(["Col A"], ["missing"], required=False) is None

    def test_case_insensitive(self):
        headers = ["EXTERNAL DOCUMENT NO."]
        assert ms.find_column(headers, ["external document no"]) == 0


class TestDetectColumns:
    def test_bc_column_hints_all_found(self):
        headers = list(BC_HEADERS)
        cols = ms.detect_columns(
            headers,
            ms.BC_COLUMN_HINTS,
            labels_required={"ext_doc_no", "rfp_no", "remaining_amount", "due_date", "open"},
        )
        assert cols["ext_doc_no"] == 0
        assert cols["rfp_no"] == 1
        assert cols["remaining_amount"] == 6  # not the sales-invoice column
        assert cols["remaining_sales_amt"] == 7
        assert cols["due_date"] == 8
        assert cols["open"] == 9

    def test_remaining_amount_does_not_grab_sales_invoice(self):
        # Put sales-invoice column BEFORE plain remaining-amount
        headers = [
            "Remaining Amount Sales Invoice",
            "Remaining Amount",
            "External Document No.",
            "RFP No.",
            "Due Date",
            "Open",
        ]
        cols = ms.detect_columns(
            headers, ms.BC_COLUMN_HINTS,
            labels_required={"ext_doc_no", "rfp_no", "remaining_amount", "due_date", "open"},
        )
        assert cols["remaining_amount"] == 1         # plain "Remaining Amount"
        assert cols["remaining_sales_amt"] == 0      # "Remaining Amount Sales Invoice"


# ===========================================================================
# 2. match_status — build_bc_dicts_from_rows
# ===========================================================================

class TestBuildBcDictsFromRows:
    def _build(self, *data_rows):
        headers, rows = _make_bc_rows(*data_rows)
        return ms.build_bc_dicts_from_rows(headers, rows)

    def test_single_row_indexed_by_ext_and_rfp(self):
        by_ext, by_rfp = self._build(BC_ROW_OPEN_FUTURE)
        assert "INV-001" in by_ext
        assert "RFP-100" in by_rfp
        assert by_ext["INV-001"]["remaining_amount"] == -500.0
        assert by_ext["INV-001"]["remaining_sales_amt"] == 600.0

    def test_multiple_rows(self):
        by_ext, by_rfp = self._build(BC_ROW_OPEN_FUTURE, BC_ROW_PAID)
        assert len(by_ext) == 2
        assert "INV-001" in by_ext and "INV-003" in by_ext

    def test_rfp_first_occurrence_wins(self):
        row_a = ("INV-A", "RFP-SAME", "P-A", "V1", "Vendor", "", -100.0, 0.0, FUTURE, True, "")
        row_b = ("INV-B", "RFP-SAME", "P-B", "V1", "Vendor", "", -200.0, 0.0, FUTURE, True, "")
        _, by_rfp = self._build(row_a, row_b)
        assert by_rfp["RFP-SAME"]["ext_doc_no"] == "INV-A"

    def test_empty_ext_doc_skipped(self):
        row = ("", "RFP-X", "P", "V", "Name", "", -50.0, 0.0, FUTURE, True, "")
        by_ext, by_rfp = self._build(row)
        assert len(by_ext) == 0
        assert "RFP-X" in by_rfp

    def test_missing_required_column_raises(self):
        with pytest.raises(ValueError):
            ms.build_bc_dicts_from_rows(("Only Col",), [("value",)])


# ===========================================================================
# 3. match_status — derive_payment_status
# ===========================================================================

class TestDerivePaymentStatus:
    def _rec(self, open_flag, due_date=None):
        return {"open": open_flag, "due_date": due_date}

    def test_none_rec_returns_not_found(self):
        assert ms.derive_payment_status(None, TODAY) == ms.STATUS_NOT_FOUND

    def test_closed_flag_returns_paid(self):
        assert ms.derive_payment_status(self._rec(False), TODAY) == ms.STATUS_PAID
        assert ms.derive_payment_status(self._rec("0"), TODAY) == ms.STATUS_PAID

    def test_open_past_due_returns_overdue(self):
        assert ms.derive_payment_status(self._rec(True, PAST), TODAY) == ms.STATUS_OVERDUE

    def test_open_future_due_returns_scheduled(self):
        assert ms.derive_payment_status(self._rec(True, FUTURE), TODAY) == ms.STATUS_SCHEDULED

    def test_open_no_due_date_returns_scheduled(self):
        assert ms.derive_payment_status(self._rec(True, None), TODAY) == ms.STATUS_SCHEDULED


# ===========================================================================
# 4. match_status — process_in_memory
# ===========================================================================

class TestProcessInMemory:
    """Integration tests for process_in_memory: matching logic + Excel output."""

    def setup_method(self):
        self.tmp_dir = tempfile.mkdtemp()
        headers, rows = _make_bc_rows(
            BC_ROW_OPEN_FUTURE,
            BC_ROW_OPEN_PAST,
            BC_ROW_PAID,
        )
        self.by_ext, self.by_rfp = ms.build_bc_dicts_from_rows(headers, rows)

    def teardown_method(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def _run(self, pairs):
        return ms.process_in_memory(pairs, self.by_ext, self.by_rfp, self.tmp_dir, today=TODAY)

    def test_exact_invoice_match(self):
        stats, headers, rows_out, path = self._run([{"invoice_no": "INV-001", "rfp_no": ""}])
        assert stats["matched"] == 1
        assert stats["not_found"] == 0
        row = rows_out[0]
        assert row[headers.index("Match status")] == "Matched"
        assert row[headers.index("Status update")] == ms.STATUS_SCHEDULED

    def test_rfp_only_match(self):
        stats, headers, rows_out, path = self._run([{"invoice_no": "", "rfp_no": "RFP-200"}])
        assert stats["matched"] == 1
        row = rows_out[0]
        assert row[headers.index("Match status")] == "Matched"

    def test_not_found(self):
        stats, headers, rows_out, path = self._run([{"invoice_no": "FAKE-999", "rfp_no": ""}])
        assert stats["not_found"] == 1
        assert rows_out[0][headers.index("Match status")] == "Not found"
        assert rows_out[0][headers.index("Status update")] == ms.STATUS_NOT_FOUND

    def test_paid_invoice(self):
        stats, headers, rows_out, path = self._run([{"invoice_no": "INV-003", "rfp_no": "RFP-300"}])
        assert stats["matched"] == 1
        assert rows_out[0][headers.index("Status update")] == ms.STATUS_PAID

    def test_overdue_invoice(self):
        _, headers, rows_out, _ = self._run([{"invoice_no": "INV-002", "rfp_no": "RFP-200"}])
        assert rows_out[0][headers.index("Status update")] == ms.STATUS_OVERDUE

    def test_discrepancy_when_invoice_and_rfp_disagree(self):
        # INV-001 belongs to RFP-100; pair it with the wrong RFP
        _, headers, rows_out, _ = self._run([{"invoice_no": "INV-001", "rfp_no": "RFP-200"}])
        assert rows_out[0][headers.index("Match status")] == "Discrepancy"

    def test_multiple_pairs_mixed(self):
        pairs = [
            {"invoice_no": "INV-001", "rfp_no": "RFP-100"},  # matched
            {"invoice_no": "FAKE", "rfp_no": ""},             # not found
            {"invoice_no": "INV-003", "rfp_no": "RFP-300"},  # matched (paid)
        ]
        stats, headers, rows_out, path = self._run(pairs)
        assert stats["matched"] == 2
        assert stats["not_found"] == 1
        assert len(rows_out) == 3

    def test_output_excel_written_and_readable(self):
        _, out_headers, _, path = self._run([{"invoice_no": "INV-001", "rfp_no": ""}])
        assert os.path.exists(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        file_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()
        assert file_headers == out_headers

    def test_sales_invoice_column_in_output(self):
        _, headers, rows_out, _ = self._run([{"invoice_no": "INV-001", "rfp_no": ""}])
        assert "BC: Remaining Amt. Sales Invoice" in headers
        col = headers.index("BC: Remaining Amt. Sales Invoice")
        assert rows_out[0][col] == 600.0

    def test_empty_pairs_returns_zero_rows(self):
        stats, _, rows_out, _ = self._run([])
        assert rows_out == []
        assert stats == {"matched": 0, "discrepancy": 0, "not_found": 0}


# ===========================================================================
# 5. payment_category — payment status rules
# ===========================================================================

class TestPaymentCategoryStatus:
    def _status(self, lcy, sales, due=None):
        return pc._derive_payment_status(lcy, sales, due, today=TODAY)

    def test_both_none_unknown(self):
        assert self._status(None, None) == pc.STATUS_UNKNOWN

    def test_lcy_zero_sales_zero_paid(self):
        assert self._status(0, 0) == pc.STATUS_PAID

    def test_lcy_zero_sales_nonzero_rare_overlap(self):
        assert self._status(0, 100) == pc.STATUS_PAID_BUT_CUSTOMER_NOT_PAID

    def test_lcy_nonzero_sales_nonzero_customer_not_paid(self):
        assert self._status(-500, 600) == pc.STATUS_CUSTOMER_NOT_PAID

    def test_ready_future_due_not_due_yet(self):
        assert self._status(-300, 0, FUTURE) == pc.STATUS_NOT_DUE_YET

    def test_ready_past_due_pay_next_cycle(self):
        assert self._status(-300, 0, PAST) == pc.STATUS_PAY_NEXT_CYCLE

    def test_ready_within_window_pay_next_cycle(self):
        within_window = TODAY + timedelta(days=1)
        assert self._status(-300, 0, within_window) == pc.STATUS_PAY_NEXT_CYCLE

    def test_ready_no_due_date_unknown_due(self):
        assert self._status(-300, 0, None) == pc.STATUS_DUE_DATE_UNKNOWN

    def test_string_numbers_parsed(self):
        assert self._status("0", "0") == pc.STATUS_PAID
        assert self._status("-300.50", "0") in (
            pc.STATUS_PAY_NEXT_CYCLE, pc.STATUS_NOT_DUE_YET, pc.STATUS_DUE_DATE_UNKNOWN
        )


# ===========================================================================
# 6. payment_category — load_ledger_from_rows
# ===========================================================================

PC_HEADERS = (
    "External Document No.",
    "Vendor Name",
    "RFP No.",
    "Remaining Amount",
    "Posting Date",
    "Due Date",
    "Entry No.",
    "Remaining Amt. (LCY)",
    "Remaining Amount Sales Invoice",
    "Open",
)


def _make_pc_rows(*data_rows):
    return PC_HEADERS, list(data_rows)


PC_ROW_A = ("INV-001", "Hotel Alpha", "RFP-100", -500.0,
             date(2026, 7, 1), FUTURE, 1, -500.0, 600.0, True)
PC_ROW_B = ("INV-002", "Hotel Beta", "RFP-200", 0.0,
             date(2026, 6, 1), PAST, 2, 0.0, 0.0, False)


class TestLoadLedgerFromRows:
    def test_builds_by_invoice_index(self):
        headers, rows = _make_pc_rows(PC_ROW_A, PC_ROW_B)
        idx = pc.load_ledger_from_rows(headers, rows)
        assert "INV-001" in idx.by_invoice
        assert "INV-002" in idx.by_invoice

    def test_builds_by_rfp_index(self):
        headers, rows = _make_pc_rows(PC_ROW_A)
        idx = pc.load_ledger_from_rows(headers, rows)
        assert "RFP-100" in idx.by_rfp

    def test_rows_list_populated(self):
        headers, rows = _make_pc_rows(PC_ROW_A, PC_ROW_B)
        idx = pc.load_ledger_from_rows(headers, rows)
        assert len(idx.rows) == 2

    def test_status_derived_at_load_time(self):
        headers, rows = _make_pc_rows(PC_ROW_A)
        idx = pc.load_ledger_from_rows(headers, rows)
        # INV-001: lcy=-500, sales=600 → Customer Not Paid
        assert idx.by_invoice["INV-001"]["status"] == pc.STATUS_CUSTOMER_NOT_PAID

    def test_empty_file_returns_empty_index(self):
        idx = pc.load_ledger_from_rows(PC_HEADERS, [])
        assert idx.rows == []
        assert idx.by_invoice == {}
        assert idx.by_rfp == {}

    def test_match_by_invoice_no(self):
        headers, rows = _make_pc_rows(PC_ROW_A)
        idx = pc.load_ledger_from_rows(headers, rows)
        matched = pc.match_by_invoice_no(idx, "inv-001")   # case-insensitive
        assert matched is not None
        assert matched["invoice_no"] == "INV-001"

    def test_match_by_invoice_no_miss(self):
        headers, rows = _make_pc_rows(PC_ROW_A)
        idx = pc.load_ledger_from_rows(headers, rows)
        assert pc.match_by_invoice_no(idx, "MISSING") is None

    def test_match_by_rfp_no(self):
        headers, rows = _make_pc_rows(PC_ROW_A)
        idx = pc.load_ledger_from_rows(headers, rows)
        matched = pc.match_by_rfp_no(idx, "rfp-100")
        assert matched is not None

    def test_match_empty_returns_none(self):
        headers, rows = _make_pc_rows(PC_ROW_A)
        idx = pc.load_ledger_from_rows(headers, rows)
        assert pc.match_by_invoice_no(idx, "") is None
        assert pc.match_by_rfp_no(idx, None) is None


# ===========================================================================
# 7. Unified cache — server._load_unified_ledger / _get_unified_ledger
# ===========================================================================

class TestUnifiedCache:
    """Tests for the server-side _UnifiedLedger cache without starting Flask."""

    def setup_method(self):
        self.tmp_dir = tempfile.mkdtemp()
        # Build a minimal BC-format xlsx on disk
        self.bc_path = os.path.join(self.tmp_dir, "bc_export.xlsx")
        rows = [BC_HEADERS] + [BC_ROW_OPEN_FUTURE, BC_ROW_PAID]
        _save_wb(rows, self.bc_path)
        # Import server and clear its cache between tests
        import server
        self.server = server
        server._UNIFIED_CACHE.clear()

    def teardown_method(self):
        self.server._UNIFIED_CACHE.clear()
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_load_unified_builds_pc_index(self):
        ledger = self.server._load_unified_ledger(self.bc_path)
        assert hasattr(ledger, "pc_index")
        assert hasattr(ledger.pc_index, "by_invoice")

    def test_load_unified_builds_by_ext_and_by_rfp(self):
        ledger = self.server._load_unified_ledger(self.bc_path)
        assert "INV-001" in ledger.by_ext
        assert "RFP-100" in ledger.by_rfp

    def test_cache_hit_returns_same_object(self):
        ledger1 = self.server._get_unified_ledger(self.bc_path)
        ledger2 = self.server._get_unified_ledger(self.bc_path)
        assert ledger1 is ledger2   # exact same object — no re-parse

    def test_cache_miss_on_mtime_change(self):
        ledger1 = self.server._get_unified_ledger(self.bc_path)
        # Force a detectable mtime change by setting it 60s in the past.
        # os.utime(path, None) can land in the same sub-second window on
        # Windows and produce an identical float → no cache miss.
        original_mtime = os.path.getmtime(self.bc_path)
        os.utime(self.bc_path, (original_mtime - 60, original_mtime - 60))
        ledger2 = self.server._get_unified_ledger(self.bc_path)
        assert ledger1 is not ledger2   # re-parsed

    def test_non_bc_file_gracefully_returns_empty_bulk_dicts(self):
        # A file with no BC columns (e.g. just "Name" and "Note")
        non_bc = os.path.join(self.tmp_dir, "non_bc.xlsx")
        _save_wb([("Name", "Note"), ("Alice", "Hi")], non_bc)
        ledger = self.server._load_unified_ledger(non_bc)
        assert ledger.by_ext == {}
        assert ledger.by_rfp == {}
        # pc_index rows list is empty (no invoice_no column)
        assert ledger.pc_index.rows == []

    def test_payment_category_and_bulk_share_same_object(self):
        """If PC warms the cache, bulk status must get the same unified object."""
        pc_view = self.server._get_unified_ledger(self.bc_path)
        bulk_view = self.server._get_unified_ledger(self.bc_path)
        assert pc_view is bulk_view


# ===========================================================================
# 8. Unified cache — Flask API integration (no Claude)
# ===========================================================================

@pytest.fixture(scope="module")
def flask_client():
    """Return a Flask test client with the cache cleared."""
    import server
    server.app.config["TESTING"] = True
    server._UNIFIED_CACHE.clear()
    with server.app.test_client() as client:
        yield client
    server._UNIFIED_CACHE.clear()


@pytest.fixture(scope="module")
def bc_xlsx(tmp_path_factory):
    """Write a minimal BC xlsx file to disk for Flask route tests."""
    d = tmp_path_factory.mktemp("bc")
    path = str(d / "bc_export.xlsx")
    rows = [BC_HEADERS] + [BC_ROW_OPEN_FUTURE, BC_ROW_OPEN_PAST, BC_ROW_PAID]
    _save_wb(rows, path)
    return path


class TestFlaskBulkStatusTableMode:
    def test_table_mode_matched(self, flask_client, bc_xlsx):
        import server; server._UNIFIED_CACHE.clear()
        resp = flask_client.post("/api/run-bulk-status", json={
            "bc_export_path": bc_xlsx,
            "table_rows": [{"invoice_no": "INV-001", "rfp_no": "RFP-100"}],
        })
        assert resp.status_code == 200
        data = resp.get_json()
        assert data.get("rows")
        assert data["rows"][0]["Match status"] == "Matched"

    def test_table_mode_not_found(self, flask_client, bc_xlsx):
        import server; server._UNIFIED_CACHE.clear()
        resp = flask_client.post("/api/run-bulk-status", json={
            "bc_export_path": bc_xlsx,
            "table_rows": [{"invoice_no": "NOPE-999", "rfp_no": ""}],
        })
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["rows"][0]["Match status"] == "Not found"

    def test_table_mode_includes_client_status_columns(self, flask_client, bc_xlsx):
        import server; server._UNIFIED_CACHE.clear()
        resp = flask_client.post("/api/run-bulk-status", json={
            "bc_export_path": bc_xlsx,
            "table_rows": [{"invoice_no": "INV-001", "rfp_no": ""}],
        })
        data = resp.get_json()
        row = data["rows"][0]
        assert "BC: Remaining Amt. Sales Invoice" in row
        assert "BC: Due Date" in row
        assert "Status update" in row

    def test_missing_bc_path_returns_400(self, flask_client):
        resp = flask_client.post("/api/run-bulk-status", json={
            "table_rows": [{"invoice_no": "INV-001", "rfp_no": ""}],
        })
        assert resp.status_code == 400

    def test_empty_table_returns_400(self, flask_client, bc_xlsx):
        resp = flask_client.post("/api/run-bulk-status", json={
            "bc_export_path": bc_xlsx,
            "table_rows": [],
        })
        assert resp.status_code == 400

    def test_second_call_uses_cache(self, flask_client, bc_xlsx):
        import server; server._UNIFIED_CACHE.clear()
        payload = {"bc_export_path": bc_xlsx,
                   "table_rows": [{"invoice_no": "INV-001", "rfp_no": ""}]}
        flask_client.post("/api/run-bulk-status", json=payload)
        assert len(server._UNIFIED_CACHE) == 1
        # Second call — cache must still have exactly 1 entry (no re-parse)
        flask_client.post("/api/run-bulk-status", json=payload)
        assert len(server._UNIFIED_CACHE) == 1


class TestFlaskPaymentCategory:
    def test_number_mode_match(self, flask_client, bc_xlsx):
        import server; server._UNIFIED_CACHE.clear()
        resp = flask_client.post("/api/payment-category/run", json={
            "input_type": "text",
            "text_mode": "number",
            "text_value": "INV-001",
            "ledger_path": bc_xlsx,
        })
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["records"][0]["Match"] == "Yes"

    def test_number_mode_miss(self, flask_client, bc_xlsx):
        import server; server._UNIFIED_CACHE.clear()
        resp = flask_client.post("/api/payment-category/run", json={
            "input_type": "text",
            "text_mode": "number",
            "text_value": "NOT-FOUND",
            "ledger_path": bc_xlsx,
        })
        assert resp.status_code == 200
        assert resp.get_json()["records"][0]["Match"] == "No"

    def test_missing_ledger_path_returns_400(self, flask_client):
        resp = flask_client.post("/api/payment-category/run", json={
            "input_type": "text",
            "text_mode": "number",
            "text_value": "INV-001",
        })
        assert resp.status_code == 400

    def test_pc_and_bulk_share_cache(self, flask_client, bc_xlsx):
        """Payment Category warming the cache must be reused by Bulk Status."""
        import server; server._UNIFIED_CACHE.clear()
        # Warm via Payment Category
        flask_client.post("/api/payment-category/run", json={
            "input_type": "text", "text_mode": "number",
            "text_value": "INV-001", "ledger_path": bc_xlsx,
        })
        assert len(server._UNIFIED_CACHE) == 1
        # Bulk Status with the SAME file must NOT add a second cache entry
        flask_client.post("/api/run-bulk-status", json={
            "bc_export_path": bc_xlsx,
            "table_rows": [{"invoice_no": "INV-001", "rfp_no": ""}],
        })
        assert len(server._UNIFIED_CACHE) == 1
